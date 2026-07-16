"""
Cong cu TUY CHON: gop du lieu tu 3 file Excel rieng cua ban cu (moi tac vu 1
file .xlsx) vao 1 file Excel gop moi (config.EXCEL_FILE, moi tac vu 1 sheet).

Chi can chay 1 LAN DUY NHAT khi chuyen tu bo 3 script rieng sang du an gop nay,
neu ban muon giu lai lich su da tong hop truoc do. Neu khong can giu lich su cu,
bo qua script nay - cac tac vu se tu tao sheet moi, trong, trong lan chay dau tien.

CACH DUNG
---------
1) Sua OLD_FILES ben duoi cho dung duong dan 3 file Excel cu tren may ban.
2) Chay: python migrate_old_excel.py
3) Kiem tra lai config.EXCEL_FILE - du lieu cu duoc them vao dung sheet tuong
   ung. Dong da trung (theo cung khoa kiem tra trung trong config.DUPLICATE_CHECK_MODE)
   se tu dong duoc BO QUA, khong ghi lap.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict

from openpyxl import load_workbook

import config
from do_auto import excel_log

# Sua lai duong dan 3 file Excel cu cho dung tren may cua ban. Neu file nao
# khong ton tai hoac khong can gop, cu de nguyen - script se tu bo qua.
OLD_FILES: Dict[str, Path] = {
    "chu_tri": Path(r"D:\OneDrive - NPT\9. Jobs\Van_ban\VB_Chu_tri_da_XL\Tong_hop_VB_chu_tri_da_XL.xlsx"),
    "phoi_hop": Path(r"D:\OneDrive - NPT\9. Jobs\Van_ban\VB_Phoi_hop\Tong_hop_VB_phoi_hop.xlsx"),
    "dang_doan": Path(r"D:\OneDrive - NPT\9. Jobs\Van_ban\VB_Dang_doan_phoi_hop\Tong_hop_VB_Dang_doan_phoi_hop.xlsx"),
}

# Cac file cu deu co cung layout: dong 1 = tieu de, dong 2 = header, dong 3 tro
# di = du lieu, thu tu cot giong het HEADERS trong do_auto/excel_log.py.
OLD_DATA_START_ROW = 3


def migrate_one(task_key: str, old_path: Path) -> None:
    task = config.TASKS.get(task_key)
    if task is None:
        print(f"⚠️ Không có tác vụ '{task_key}' trong config.TASKS, bỏ qua.")
        return
    if not old_path.exists():
        print(f"⏸️ Không thấy file cũ cho '{task_key}': {old_path} - bỏ qua.")
        return

    print(f"\n--- Gộp dữ liệu cũ cho tác vụ: {task.label} ---")
    print(f"Nguồn: {old_path}")

    wb_old = load_workbook(old_path, read_only=True, data_only=True)
    ws_old = wb_old.active

    existing_keys = excel_log.load_existing_duplicate_keys(config.EXCEL_FILE, task.sheet_name, config.DUPLICATE_CHECK_MODE)

    added = 0
    skipped = 0
    for row in ws_old.iter_rows(min_row=OLD_DATA_START_ROW, values_only=True):
        if not row or len(row) < 13:
            continue

        data = {
            "so_vb": str(row[1] or ""),
            "ngay_vb": str(row[2] or ""),
            "noi_phat_hanh": str(row[3] or ""),
            "trich_yeu": str(row[4] or ""),
            "nguoi_chi_dao": str(row[5] or ""),
            "thoi_gian_chi_dao": str(row[6] or ""),
            "noi_dung_chi_dao": str(row[7] or ""),
            "chu_tri": str(row[8] or ""),
            "phoi_hop": str(row[9] or ""),
            "thu_muc_luu": str(row[10] or ""),
            "ten_file_luu": str(row[11] or ""),
            "thoi_gian_luu": str(row[12] or ""),
        }
        if not data["so_vb"].strip():
            continue

        key = excel_log.build_duplicate_key(data, config.DUPLICATE_CHECK_MODE)
        if key in existing_keys:
            skipped += 1
            continue

        excel_log.append_excel_log(config.EXCEL_FILE, task.sheet_name, task.title_text, data)
        existing_keys.add(key)
        added += 1

    wb_old.close()
    print(f"✅ Xong '{task.label}': thêm mới {added} dòng, bỏ qua {skipped} dòng đã trùng.")


def main() -> None:
    excel_log.ensure_all_sheets(config.EXCEL_FILE, config.TASKS)
    for task_key, old_path in OLD_FILES.items():
        migrate_one(task_key, old_path)
    print(f"\n📘 Hoàn tất. File gộp: {config.EXCEL_FILE}")


if __name__ == "__main__":
    main()
