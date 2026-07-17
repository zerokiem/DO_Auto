"""
Ghi log Excel gop - 1 file .xlsx, moi tac vu 1 sheet rieng
(vd: Chu_tri, Phoi_hop, Dang_doan).

Luu y quan trong so voi 3 script goc: 3 script goc moi script dung 1 file .xlsx
rieng nen ham init luon dung wb.active la an toan. Khi gop chung 1 file nhieu
sheet, KHONG duoc dung wb.active de tim/doi ten sheet nua (se lam hong sheet
dang active neu sheet muc tieu chua ton tai) - ham init_excel_log() ben duoi da
sua loi nay: neu sheet can dung chua co, se tao sheet moi (create_sheet) thay vi
doi ten sheet dang active.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Set
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import text_utils

HEADERS = [
    "STT",
    "Số VB",
    "Ngày VB",
    "Nơi phát hành",
    "Trích yếu",
    "Người chỉ đạo",
    "Thời gian chỉ đạo",
    "Nội dung chỉ đạo",
    "Chủ trì",
    "Phối hợp",
    "Thư mục lưu",
    "Tên file lưu",
    "Thời gian lưu",
]

TITLE_ROW = 1
HEADER_ROW = 2
DATA_START_ROW = 3

# Do rong cot (don vi "ky tu" - dung ca cho Excel column width lan cho web, xem
# webapp/app.py excel_page() va webapp/templates/excel_view.html). La NGUON DUY
# NHAT cho do rong cot, tranh viet 2 lan 2 noi roi bi lech nhau.
COLUMN_WIDTHS = dict(
    zip(
        HEADERS,
        [6, 16.6, 13, 30, 50, 15.5, 20, 30, 11.1, 30, 22.2, 45, 20],
    )
)

# Ten sheet kieu cu (truoc khi doi sang tieng Viet co dau) - de tu dong doi ten
# sheet cu thanh ten moi thay vi tao sheet rong moi lam "mo côi" du lieu cu.
LEGACY_SHEET_NAMES = {
    "Chủ trì": ["Chu_tri"],
    "Phối hợp": ["Phoi_hop"],
    "Đảng - Đoàn": ["Dang_doan"],
}

# Cot theo chi so 0-based trong openpyxl values_only iterator (0 = STT).
COL_SO_VB = 1
COL_NGAY_VB = 2
COL_THOI_GIAN_CHI_DAO = 6
COL_TEN_FILE_LUU = 11


def apply_excel_layout(ws, title_text: str) -> None:
    """Chuan hoa layout Excel: title, header, do rong cot, freeze, filter, an cot."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(color="1F4E78", bold=True, size=14)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    first_row_values = [ws.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)]
    if first_row_values == HEADERS:
        ws.insert_rows(1)

    ws.cell(row=TITLE_ROW, column=1).value = title_text
    try:
        for merged in list(ws.merged_cells.ranges):
            if str(merged).startswith("A1:"):
                ws.unmerge_cells(str(merged))
    except Exception:
        pass

    title_cell = ws.cell(row=TITLE_ROW, column=1)
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = border
    ws.row_dimensions[TITLE_ROW].height = 28

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[HEADER_ROW].height = 35

    widths = {i: COLUMN_WIDTHS[header] for i, header in enumerate(HEADERS, start=1)}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # An cot G = Thoi gian chi dao. Can thi Unhide lai trong Excel.
    ws.column_dimensions["G"].hidden = True

    ws.freeze_panes = "D3"
    ws.auto_filter.ref = "A2:M2"

    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, max_col=len(HEADERS)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _get_or_create_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    # Ten sheet vua doi sang tieng Viet co dau (vd "Chu_tri" -> "Chủ trì")? Doi
    # ten sheet cu thay vi tao sheet moi rong, de khong "mo côi" du lieu da co.
    for legacy_name in LEGACY_SHEET_NAMES.get(sheet_name, []):
        if legacy_name in wb.sheetnames:
            ws = wb[legacy_name]
            ws.title = sheet_name
            return ws
    # Neu workbook chi co dung 1 sheet mac dinh trong ("Sheet") thi doi ten sheet do
    # thay vi tao them sheet rong thua.
    if wb.sheetnames == ["Sheet"] and wb["Sheet"].max_row <= 1:
        ws = wb["Sheet"]
        ws.title = sheet_name
        return ws
    return wb.create_sheet(title=sheet_name)


def init_excel_log(excel_file: Path, sheet_name: str, title_text: str) -> None:
    """Dam bao file Excel va sheet muc tieu ton tai, dung layout. An toan khi goi
    nhieu lan / cho nhieu sheet khac nhau trong cung 1 workbook."""
    excel_file.parent.mkdir(parents=True, exist_ok=True)

    if excel_file.exists():
        wb = load_workbook(excel_file)
        ws = _get_or_create_sheet(wb, sheet_name)
        apply_excel_layout(ws, title_text)
        wb.save(excel_file)
        wb.close()
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    apply_excel_layout(ws, title_text)
    wb.save(excel_file)
    wb.close()
    print(f"✅ Đã tạo file Excel tổng hợp: {excel_file}")


def ensure_all_sheets(excel_file: Path, tasks: Dict[str, "object"]) -> None:
    """Tao san tat ca sheet cho moi tac vu trong TASKS theo dung thu tu sheet_order,
    de thu tu tab trong Excel luon nhat quan du nguoi dung chay tac vu nao truoc."""
    excel_file.parent.mkdir(parents=True, exist_ok=True)
    ordered = sorted(tasks.values(), key=lambda t: t.sheet_order)

    if excel_file.exists():
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    for task in ordered:
        ws = _get_or_create_sheet(wb, task.sheet_name)
        apply_excel_layout(ws, task.title_text)

    wb.save(excel_file)
    wb.close()


def get_next_stt(ws) -> int:
    if ws.max_row < DATA_START_ROW:
        return 1
    return max(1, ws.max_row - HEADER_ROW + 1)


def get_next_excel_stt(excel_file: Path, sheet_name: str) -> int:
    init_excel_log(excel_file, sheet_name, sheet_name)
    wb = load_workbook(excel_file, read_only=False)
    ws = wb[sheet_name]
    stt = get_next_stt(ws)
    wb.close()
    return stt


def file_uri_from_path(path_text: str) -> str:
    """Tao URI file:/// tu duong dan (Windows hoac POSIX) de nhung vao Excel.
    KHONG dung Path.resolve(): khi chay trong container (Linux) ma duong dan la
    kieu Windows (vd 'S:\\...') thi resolve() se hieu nham thanh duong dan tuong
    doi va noi them cwd vao -> hong link. O day chi chuyen '\\' -> '/', escape
    dau cach, roi them tien to file:// phu hop."""
    normalized = path_text.replace("\\", "/").replace(" ", "%20")
    if normalized.startswith("/"):
        return "file://" + normalized  # POSIX tuyet doi: /data/... -> file:///data/...
    return "file:///" + normalized  # Windows co o dia: S:/... -> file:///S:/...


def to_display_folder(actual_folder, download_base, display_base) -> str:
    """Doi duong dan THUC TE (noi file that su duoc luu, vd '/data/VB_Chu_tri_da_XL'
    trong container) sang duong dan HIEN THI de nguoi dung bam mo tren Windows
    (vd 'S:\\Working\\Van_ban\\VB_Chu_tri_da_XL').

    Neu display_base trung download_base (dang chay truc tiep tren Windows, khong
    dat DOFFICE_DISPLAY_DIR) thi tra ve nguyen ban - khong doi gi."""
    actual = Path(actual_folder)
    base = Path(download_base)
    display_base = str(display_base).rstrip("/\\")
    if str(base) == display_base:
        return str(actual)
    try:
        rel = actual.relative_to(base)
    except ValueError:
        return str(actual)  # khong nam duoi base -> khong biet map, giu nguyen
    if str(rel) in (".", ""):
        return display_base
    return display_base + "\\" + str(rel).replace("/", "\\")


def subdir_of(folder: str) -> str:
    """Lay thanh phan cuoi cua duong dan thu muc (ten thu muc con), du duong dan
    la kieu Windows ('S:\\..\\VB_Chu_tri_da_XL') hay POSIX ('/data/VB_Chu_tri_da_XL')."""
    parts = [p for p in re.split(r"[\\/]+", str(folder).strip()) if p and p not in (".", "..")]
    return parts[-1] if parts else ""


def build_file_link(folder: str, filename: str, link_base: str) -> str:
    """Tao gia tri hyperlink cho cot 'Ten file luu'.

    - Neu link_base la URL web (http/https, vd 'http://100.100.1.254:8877/vb') thi
      tra ve URL toi NAS qua web: link_base/<thu_muc_con>/<ten_file>. Bam mo duoc
      tren dien thoai co Tailscale va moi thiet bi trong mang.
    - Nguoc lai (link_base rong) thi tra ve link file:// tren duong dan hien thi
      (o S:) nhu truoc."""
    if link_base and (link_base.startswith("http://") or link_base.startswith("https://")):
        segs = []
        sub = subdir_of(folder)
        if sub:
            segs.append(quote(sub, safe=""))
        segs.append(quote(filename, safe=""))
        return link_base.rstrip("/") + "/" + "/".join(segs)
    sep = "\\" if ("\\" in folder or (len(folder) >= 2 and folder[1] == ":")) else "/"
    return file_uri_from_path(folder.rstrip("/\\") + sep + filename)


def append_excel_log(excel_file: Path, sheet_name: str, title_text: str, data: Dict[str, str], file_link_base: str = "") -> None:
    init_excel_log(excel_file, sheet_name, title_text)

    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    stt = get_next_stt(ws)

    row = [
        stt,
        data.get("so_vb", ""),
        data.get("ngay_vb", ""),
        data.get("noi_phat_hanh", ""),
        data.get("trich_yeu", ""),
        data.get("nguoi_chi_dao", ""),
        data.get("thoi_gian_chi_dao", ""),
        data.get("noi_dung_chi_dao", ""),
        data.get("chu_tri", ""),
        data.get("phoi_hop", ""),
        data.get("thu_muc_luu", ""),
        data.get("ten_file_luu", ""),
        data.get("thoi_gian_luu", ""),
    ]
    ws.append(row)

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[ws.max_row]:
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    saved_folder = data.get("thu_muc_luu", "")
    saved_name = data.get("ten_file_luu", "")
    if saved_folder and saved_name:
        file_cell = ws.cell(row=ws.max_row, column=12)
        # Link tro toi NAS qua web (neu co DISPLAY_BASE_URL) de mo duoc tren dien
        # thoai/Tailscale; nguoc lai la link file:// tren o S:. Xem build_file_link.
        file_cell.hyperlink = build_file_link(saved_folder, saved_name, file_link_base)
        file_cell.style = "Hyperlink"
        file_cell.comment = None

    apply_excel_layout(ws, title_text)
    wb.save(excel_file)
    wb.close()
    print(f"✅ [{sheet_name}] Đã ghi Excel dòng STT {stt}: {data.get('so_vb', '')}")


def build_duplicate_key(data: Dict[str, str], duplicate_check_mode: str) -> str:
    so_vb = text_utils.normalize_key(data.get("so_vb", ""))
    ngay_vb = text_utils.normalize_key(data.get("ngay_vb", ""))
    if duplicate_check_mode == "so_vb_time":
        tg = text_utils.normalize_key(data.get("thoi_gian_chi_dao", ""))
        return f"{so_vb}|{tg}"
    if duplicate_check_mode == "so_vb":
        return so_vb
    return f"{so_vb}|{ngay_vb}"


def load_existing_duplicate_keys(excel_file: Path, sheet_name: str, duplicate_check_mode: str) -> Set[str]:
    init_excel_log(excel_file, sheet_name, sheet_name)
    keys: Set[str] = set()

    wb = load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row or len(row) < 7:
            continue
        so_vb = str(row[COL_SO_VB] or "")
        ngay_vb = str(row[COL_NGAY_VB] or "")
        thoi_gian = str(row[COL_THOI_GIAN_CHI_DAO] or "")
        if not so_vb.strip():
            continue
        if duplicate_check_mode == "so_vb_time":
            keys.add(f"{text_utils.normalize_key(so_vb)}|{text_utils.normalize_key(thoi_gian)}")
        elif duplicate_check_mode == "so_vb":
            keys.add(text_utils.normalize_key(so_vb))
        else:
            keys.add(f"{text_utils.normalize_key(so_vb)}|{text_utils.normalize_key(ngay_vb)}")

    wb.close()
    return keys


def load_existing_filenames(excel_file: Path, sheet_name: str) -> Set[str]:
    init_excel_log(excel_file, sheet_name, sheet_name)
    names: Set[str] = set()
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        filename = str(row[COL_TEN_FILE_LUU] or "") if row and len(row) > COL_TEN_FILE_LUU else ""
        if filename.strip():
            names.add(text_utils.normalize_key(filename))
    wb.close()
    return names
