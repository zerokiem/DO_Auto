"""Regression tests for configurable navigation and dynamic tasks."""
from __future__ import annotations

import ast
import io
import shutil
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from do_auto import browser_nav, excel_log, extract, finish_doc, settings_store
from do_auto.task_types import TaskConfig
from webapp.app import _dashboard_file_link, app


class DynamicTaskSettingsTests(unittest.TestCase):
    def test_dashboard_file_link_uses_internal_pdf_route_for_local_files(self):
        base = Path("C:/doffice-data")
        cfg = SimpleNamespace(DOWNLOAD_BASE_DIR=base, DISPLAY_BASE_URL="")
        with app.test_request_context():
            link = _dashboard_file_link(str(base / "VB_phoi_hop"), "van-ban.pdf", cfg)
        self.assertEqual(link, "/vb/VB_phoi_hop/van-ban.pdf")

    def test_download_dir_override_rebases_all_default_output_files(self):
        old_base = Path("C:/old-doffice-data")
        new_base = Path("C:/new-doffice-data")
        base_cfg = SimpleNamespace(
            AUTH_STATE=Path("playwright/.auth/state.json"),
            DOWNLOAD_BASE_DIR=old_base,
            DOWNLOAD_BASE_DIR_OVERRIDE=str(new_base),
            DISPLAY_BASE_DIR=r"S:\Mapped-but-not-actual",
            EXCEL_FILE=old_base / "Tong_hop_DOffice.xlsx",
            HISTORY_DB=old_base / "doffice_auto_history.jsonl",
            LOGS_DIR=old_base / "logs",
            TASKS={},
        )

        effective = settings_store.build_effective_config(base_cfg)

        self.assertEqual(effective.DOWNLOAD_BASE_DIR, new_base)
        self.assertEqual(effective.DISPLAY_BASE_DIR, str(new_base))
        self.assertEqual(effective.EXCEL_FILE, new_base / "Tong_hop_DOffice.xlsx")
        self.assertEqual(effective.HISTORY_DB, new_base / "doffice_auto_history.jsonl")
        self.assertEqual(effective.LOGS_DIR, new_base / "logs")

    def test_download_dir_override_can_be_written_to_older_config(self):
        original_path = settings_store.CONFIG_PATH
        with tempfile.TemporaryDirectory() as temp_dir:
            config_copy = Path(temp_dir) / "config.py"
            shutil.copy2(original_path, config_copy)
            settings_store.CONFIG_PATH = config_copy
            try:
                settings_store.update_common_fields(
                    {"DOWNLOAD_BASE_DIR_OVERRIDE": r"D:\DOffice_Data"}
                )
                source = config_copy.read_text(encoding="utf-8")
                self.assertIn('DOWNLOAD_BASE_DIR_OVERRIDE = "D:\\\\DOffice_Data"', source)
                ast.parse(source)
            finally:
                settings_store.CONFIG_PATH = original_path

    def test_open_finish_menu_clicks_ellipsis_before_menuitem(self):
        class Locator:
            def __init__(self, visible=False, on_click=None):
                self.visible = visible
                self.on_click = on_click
                self.clicked = 0

            @property
            def last(self):
                return self

            def filter(self, **_kwargs):
                return self

            def wait_for(self, state, timeout):
                if state != "visible" or not self.visible:
                    raise RuntimeError("not visible")

            def click(self, timeout):
                self.clicked += 1
                if self.on_click:
                    self.on_click()

        class Page:
            def __init__(self):
                self.menu_item = Locator()
                self.trigger = Locator(visible=True, on_click=lambda: setattr(self.menu_item, "visible", True))

            def get_by_role(self, role, name):
                return self.menu_item if role == "menuitem" else self.trigger

            def get_by_text(self, text, exact=True):
                return self.trigger

            def locator(self, selector):
                return self.trigger

        page = Page()
        with redirect_stdout(io.StringIO()):
            self.assertTrue(finish_doc.open_finish_menu(page))
        self.assertEqual(page.trigger.clicked, 1)
        self.assertTrue(page.menu_item.visible)

    def test_work_list_metadata_stays_in_four_separate_fields(self):
        class WorkRow:
            def evaluate(self, script):
                if "Boolean(row.querySelector" in script:
                    return True
                # Regression guard for the dedicated Xem de biet selectors.
                required_selectors = (
                    "td.mat-column-KY_HIEU_CV",
                    ":scope > span:nth-of-type(1)",
                    ":scope > .dokhan > span",
                    ":scope > span.text-blue-600",
                    "const workLayout",
                )
                if not all(selector in script for selector in required_selectors):
                    raise AssertionError("Missing dedicated Xem de biet selectors")
                return {
                    "so_vb": "1979/TTDTPHCM2-KHKT",
                    "ngay_vb": "14/08/2026",
                    "noi_phat_hanh": "Truyen tai dien Tp. Ho Chi Minh 2",
                    "trich_yeu": "V/v lap PATC&PBAT cho cong tac sua chua",
                    # This intentionally contains all four values. The result
                    # must still use the four targeted fields above.
                    "row_text": "1979/TTDTPHCM2-KHKT 14/08/2026 Truyen tai dien Tp. Ho Chi Minh 2 V/v lap PATC&PBAT cho cong tac sua chua",
                    "chi_dao_text": "",
                }

        row = WorkRow()
        data = extract.extract_document_info_from_row(row, mode="directive")
        self.assertEqual(data["so_vb"], "1979/TTDTPHCM2-KHKT")
        self.assertEqual(data["ngay_vb"], "14/08/2026")
        self.assertEqual(data["noi_phat_hanh"], "Truyen tai dien Tp. Ho Chi Minh 2")
        self.assertEqual(data["trich_yeu"], "V/v lap PATC&PBAT cho cong tac sua chua")

    def test_work_list_overrides_wrong_published_mode(self):
        class WorkRow:
            def evaluate(self, script):
                if "Boolean(row.querySelector" in script:
                    return True
                return {
                    "so_vb": "84/TTr-TTDTPHCM2",
                    "ngay_vb": "14/08/2026",
                    "noi_phat_hanh": "Truyen tai dien Tp. Ho Chi Minh 2",
                    "trich_yeu": "To trinh ve viec de nghi chi Quy Tuong tro xa hoi",
                    "chi_dao_text": "",
                    "row_text": "Sai fallback phai khong duoc dung",
                }

        with redirect_stdout(io.StringIO()):
            data = extract.extract_document_info_from_row(WorkRow(), mode="published")
        self.assertEqual(data["so_vb"], "84/TTr-TTDTPHCM2")
        self.assertEqual(data["ngay_vb"], "14/08/2026")
        self.assertEqual(data["noi_phat_hanh"], "Truyen tai dien Tp. Ho Chi Minh 2")
        self.assertEqual(data["trich_yeu"], "To trinh ve viec de nghi chi Quy Tuong tro xa hoi")

    def test_work_row_clicks_document_number_not_cell_center(self):
        class Page:
            def __init__(self):
                self.waits = []
                self.document_detail_visible = True

            def wait_for_timeout(self, milliseconds):
                self.waits.append(milliseconds)

        class Locator:
            def __init__(self, page, name, count=1):
                self.page = page
                self.name = name
                self._count = count
                self.clicked = 0

            @property
            def first(self):
                return self

            def count(self):
                return self._count

            def locator(self, selector):
                if selector == "div.vb-item":
                    return Locator(self.page, selector, count=0)
                return Locator(self.page, selector)

            def scroll_into_view_if_needed(self, timeout):
                self.scroll_timeout = timeout

            def click(self, timeout):
                self.clicked += 1
                self.click_timeout = timeout

        page = Page()
        row = Locator(page, "row")
        number_locator = Locator(page, "document-number")
        original_locator = row.locator

        def locator(selector):
            if "td.mat-cell > div:first-child > span:first-child" in selector:
                return number_locator
            return original_locator(selector)

        row.locator = locator
        with redirect_stdout(io.StringIO()):
            result = browser_nav.click_document_row(row, extract_mode="directive")
        self.assertTrue(result)
        self.assertEqual(number_locator.clicked, 1)

    def test_selector_step_uses_configured_locator_and_action(self):
        class Locator:
            def __init__(self):
                self.clicked = 0
                self.waited = []

            def nth(self, _index):
                return self

            def click(self, timeout):
                self.clicked += 1
                self.click_timeout = timeout

            def wait_for(self, state, timeout):
                self.waited.append((state, timeout))

        class Page:
            def __init__(self):
                self.locators = {}
                self.waits = []

            def locator(self, selector):
                return self.locators.setdefault(selector, Locator())

            def wait_for_timeout(self, milliseconds):
                self.waits.append(milliseconds)

        page = Page()
        task = TaskConfig(key="kiem_thu", label="Kiểm thử")
        ok = browser_nav.run_navigation_steps(
            page,
            task,
        )
        self.assertTrue(ok)  # [] is a valid legacy-compatible sequence.

        task.navigation_steps = [
            {"name": "Mở bộ lọc", "type": "selector", "selector": "#filter", "action": "click", "delay_ms": 0},
            {"name": "Chờ dòng", "type": "selector", "selector": "tr.row", "action": "wait", "wait_for": "visible", "delay_ms": 0},
        ]
        # Entry points normalize the Windows console to UTF-8. Capture status
        # output here so this unit test is independent of the test runner's
        # inherited code page (for example CP1252 in a piped PowerShell run).
        with redirect_stdout(io.StringIO()):
            result = browser_nav.run_navigation_steps(page, task)
        self.assertTrue(result)
        self.assertEqual(page.locators["#filter"].clicked, 1)
        self.assertEqual(page.locators["tr.row"].waited, [("visible", 10000)])

    def test_add_update_remove_keeps_config_valid(self):
        original_path = settings_store.CONFIG_PATH
        with tempfile.TemporaryDirectory() as temp_dir:
            config_copy = Path(temp_dir) / "config.py"
            shutil.copy2(original_path, config_copy)
            settings_store.CONFIG_PATH = config_copy
            try:
                settings_store.add_task(
                    {
                        "key": "kiem_thu",
                        "label": "Kiểm thử",
                        "sheet_name": "Kiểm thử",
                        "sidebar_item": "Công việc",
                        "list_link": "Đã giao việc",
                        "tab_name": "Chủ trì",
                        "navigation_steps": [],
                        "use_advanced_navigation": False,
                    }
                )
                settings_store.update_task_fields(
                    "kiem_thu",
                    {
                        "navigation_steps": [
                            {
                                "name": "Chờ dữ liệu",
                                "type": "selector",
                                "selector": "tr.row",
                                "action": "wait",
                                "empty_if_zero": True,
                            }
                        ],
                        "document_row_selector": "tr.row",
                    },
                )
                interim_source = config_copy.read_text(encoding="utf-8")
                self.assertIn('sidebar_item="Công việc"', interim_source)
                self.assertIn('list_link="Đã giao việc"', interim_source)
                self.assertIn('tab_name="Chủ trì"', interim_source)
                tree = ast.parse(interim_source)
                tasks_node = next(
                    node.value
                    for node in tree.body
                    if isinstance(node, ast.Assign)
                    and any(isinstance(target, ast.Name) and target.id == "TASKS" for target in node.targets)
                )
                task_call = next(
                    value
                    for key, value in zip(tasks_node.keys, tasks_node.values)
                    if isinstance(key, ast.Constant) and key.value == "kiem_thu"
                )
                steps_node = next(keyword.value for keyword in task_call.keywords if keyword.arg == "navigation_steps")
                self.assertTrue(ast.literal_eval(steps_node)[0]["empty_if_zero"])
                settings_store.remove_task("kiem_thu")
                final_source = config_copy.read_text(encoding="utf-8")
                ast.parse(final_source)
                self.assertNotIn('"kiem_thu": TaskConfig(', final_source)
                self.assertIn('"chu_tri": TaskConfig(', final_source)
            finally:
                settings_store.CONFIG_PATH = original_path

    def test_ensure_all_sheets_never_removes_existing_sheet(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_file = Path(temp_dir) / "tong_hop.xlsx"
            workbook = Workbook()
            workbook.active.title = "Sheet cũ"
            workbook.save(excel_file)
            workbook.close()

            excel_log.ensure_all_sheets(
                excel_file,
                {"moi": TaskConfig(key="moi", label="Mới", sheet_name="Sheet mới", title_text="TỔNG HỢP MỚI")},
            )
            workbook = load_workbook(excel_file, read_only=True)
            self.assertIn("Sheet cũ", workbook.sheetnames)
            self.assertIn("Sheet mới", workbook.sheetnames)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
