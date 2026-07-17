"""Sua loi du lieu CU cua sheet "Phoi hop": cac dong duoc ghi TRUOC KHI co
download_subdir rieng cho tung tac vu bi luu THIEU thu muc con "VB_Phoi_hop"
trong cot "Thu muc luu" (cot K) - vi du ghi "...\\Van_ban" thay vi
"...\\Van_ban\\VB_Phoi_hop". File PDF thuc te van luon nam trong VB_Phoi_hop
(da kiem tra thuc te tren NAS, khong phai o thu muc goc) - chi METADATA trong
Excel bi thieu.

Sua ca 2 noi:
  - Cot K (Thu muc luu): them "VB_Phoi_hop" vao cuoi neu chua co.
  - Cot L (Ten file luu): xay lai HYPERLINK theo cot K da sua + DISPLAY_BASE_URL.

Chay TRONG container:
    sudo -n /usr/local/bin/docker exec doffice python /app/migrate_fix_phoihop_folder.py
"""
from __future__ import annotations

import shutil
from datetime import datetime

from openpyxl import load_workbook

import config
from do_auto import excel_log

EXCEL = config.EXCEL_FILE
LINK_BASE = str(getattr(config, "DISPLAY_BASE_URL", "") or "")
SHEET = "Phối hợp"
SUBDIR = "VB_Phoi_hop"

FOLDER_COL = 11  # "Thu muc luu"
FILE_COL = 12    # "Ten file luu" (chua hyperlink)


def needs_fix(folder: str) -> bool:
    f = (folder or "").strip().replace("/", "\\").rstrip("\\").lower()
    return bool(f) and not f.endswith("\\" + SUBDIR.lower()) and not f.endswith(SUBDIR.lower())


def fixed_folder(folder: str) -> str:
    sep = "\\" if ("\\" in folder or (len(folder) >= 2 and folder[1] == ":")) else "/"
    return folder.rstrip("/\\") + sep + SUBDIR


def main() -> None:
    print(f"EXCEL     = {EXCEL}")
    print(f"LINK_BASE = {LINK_BASE!r}")
    if not EXCEL.exists():
        print("Khong thay file Excel:", EXCEL)
        return

    backup = EXCEL.with_name(EXCEL.stem + f".backup_phoihop_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(EXCEL, backup)
    print("Da sao luu:", backup)

    wb = load_workbook(EXCEL)
    if SHEET not in wb.sheetnames:
        print(f"Khong co sheet '{SHEET}'.")
        return
    ws = wb[SHEET]

    fixed = 0
    for r in range(excel_log.DATA_START_ROW, ws.max_row + 1):
        folder_cell = ws.cell(r, FOLDER_COL)
        folder = str(folder_cell.value or "")
        if not needs_fix(folder):
            continue
        new_folder = fixed_folder(folder)
        folder_cell.value = new_folder

        file_cell = ws.cell(r, FILE_COL)
        fname = str(file_cell.value or "")
        if fname:
            file_cell.hyperlink = excel_log.build_file_link(new_folder, fname, LINK_BASE)
            file_cell.style = "Hyperlink"
        fixed += 1

    wb.save(EXCEL)
    wb.close()
    print(f"XONG. Da sua {fixed} dong trong sheet '{SHEET}' (them '{SUBDIR}' vao cot K + xay lai hyperlink cot L).")


if __name__ == "__main__":
    main()
