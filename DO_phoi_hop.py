import re
import time
import unicodedata
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional, Set

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==================================================
# CẤU HÌNH
# ==================================================

# Trang DOffice. Dùng trang gốc rồi điều hướng bằng menu để ổn định hơn khi DOffice đổi route.
DOFFICE_URL = "https://doffice.npt.com.vn/"

# Session Playwright đã lưu sau khi chạy login_save_state.py
AUTH_STATE = Path("playwright/.auth/state.json")

# Thư mục lưu PDF và file Excel tổng hợp
DOWNLOAD_DIR = Path(r"D:\OneDrive - NPT\9. Jobs\Van_ban\VB_Phoi_hop")
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

LOG_FILE = DOWNLOAD_DIR / "Tong_hop_VB_phoi_hop.xlsx"

# Giới hạn số văn bản xử lý trong một lần chạy
MAX_DOCUMENTS = 147

# Có tải PDF hay không. Nếu chỉ muốn cập nhật Excel metadata thì để False.
ENABLE_DOWNLOAD_PDF = True

# Khi gặp văn bản đã có trong Excel thì dừng chương trình.
# Phù hợp khi danh sách Phối hợp sắp xếp mới nhất ở trên.
STOP_WHEN_DUPLICATE_FOUND = True

# Khóa kiểm tra trùng:
# - "so_vb_ngay_vb": Số VB + Ngày VB. Khuyến nghị dùng.
# - "so_vb": chỉ dùng Số VB.
# - "so_vb_time": Số VB + Thời gian chỉ đạo.
DUPLICATE_CHECK_MODE = "so_vb_ngay_vb"

# Sau mỗi REFRESH_LIST_EVERY văn bản thì load lại danh sách.
# Nếu ENABLE_FINISH_DOCUMENT=True thì thường không cần reload vì VB kế tiếp tự trồi lên đầu.
REFRESH_LIST_EVERY = 0  # 0 = không tự reload

# Kết thúc văn bản
ENABLE_FINISH_DOCUMENT = True
ASK_CONFIRM_BEFORE_FINISH = False

# Cách chọn văn bản kế tiếp:
# - True: sau khi Kết thúc VB, DOffice tự đẩy VB kế tiếp lên dòng đầu tiên.
# - False: khi chưa Kết thúc VB, chọn xuống dòng kế tiếp để không xử lý lại văn bản cũ.
ALWAYS_PICK_FIRST_ROW_AFTER_FINISH = True

# Tốc độ thao tác để dễ quan sát
SLOW_MO_MS = 250

# Có dừng chờ Enter trước khi đóng browser không
PAUSE_BEFORE_CLOSE = False

# Timeout bắt download theo từng lần thử.
# Lần 1 cố ý ngắn vì DOffice/PDF viewer đôi khi click đầu chỉ kích hoạt viewer, chưa phát download event.
DOWNLOAD_ATTEMPT_TIMEOUTS_MS = [7000, 15000, 20000]


# ==================================================
# EXCEL LOG
# ==================================================

HEADERS = [
    "STT",
    "Số VB",
    "Ngày VB",
    "Nơi phát hành",
    "Trích yếu",
    "Người chỉ đạo",
    "Thời gian chỉ đạo",
    "Nội dung chỉ đạo",
    "Chủ trì",
    "Phối hợp",
    "Thư mục lưu",
    "Tên file lưu",
    "Thời gian lưu",
]

TITLE_TEXT = "TỔNG HỢP CÁC VĂN BẢN PHỐI HỢP ĐÃ XỬ LÝ"
TITLE_ROW = 1
HEADER_ROW = 2
DATA_START_ROW = 3
SHEET_NAME = "Phoi_hop"


def apply_excel_layout(ws) -> None:
    """Chuẩn hóa layout Excel: title, header, width, freeze, filter."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(color="1F4E78", bold=True, size=14)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Nếu file cũ đang có header ở dòng 1 thì chèn thêm dòng tiêu đề lên trên.
    first_row_values = [ws.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)]
    if first_row_values == HEADERS:
        ws.insert_rows(1)

    # Tiêu đề lớn, ghi từ A1, không merge để dễ xử lý Excel.
    ws.cell(row=TITLE_ROW, column=1).value = TITLE_TEXT
    try:
        for merged in list(ws.merged_cells.ranges):
            if str(merged) == "A1:M1":
                ws.unmerge_cells(str(merged))
    except Exception:
        pass

    title_cell = ws.cell(row=TITLE_ROW, column=1)
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = border
    ws.row_dimensions[TITLE_ROW].height = 28

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[HEADER_ROW].height = 35

    widths = {
        1: 6,
        2: 16,
        3: 13,
        4: 30,
        5: 50,
        6: 18,
        7: 20,
        8: 30,
        9: 16,
        10: 30,
        11: 35,
        12: 45,
        13: 20,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Ẩn cột G = Thời gian chỉ đạo. Khi cần có thể Unhide trong Excel.
    ws.column_dimensions["G"].hidden = True

    ws.freeze_panes = "D3"
    ws.auto_filter.ref = "A2:M2"

    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, max_col=len(HEADERS)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def init_excel_log() -> None:
    if LOG_FILE.exists():
        wb = load_workbook(LOG_FILE)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        if ws.title != SHEET_NAME:
            ws.title = SHEET_NAME
        apply_excel_layout(ws)
        wb.save(LOG_FILE)
        wb.close()
        return

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    apply_excel_layout(ws)
    wb.save(LOG_FILE)
    wb.close()
    print(f"✅ Đã tạo file Excel tổng hợp: {LOG_FILE}")


def get_next_stt(ws) -> int:
    if ws.max_row < DATA_START_ROW:
        return 1
    return max(1, ws.max_row - HEADER_ROW + 1)


def file_uri_from_path(path_text: str) -> str:
    try:
        return Path(path_text).resolve().as_uri()
    except Exception:
        return "file:///" + path_text.replace("\\", "/").replace(" ", "%20")


def append_excel_log(data: Dict[str, str]) -> None:
    init_excel_log()

    wb = load_workbook(LOG_FILE)
    ws = wb[SHEET_NAME]
    stt = get_next_stt(ws)

    row = [
        stt,
        data.get("so_vb", ""),
        data.get("ngay_vb", ""),
        data.get("noi_phat_hanh", ""),
        data.get("trich_yeu", ""),
        data.get("nguoi_chi_dao", ""),
        data.get("thoi_gian_chi_dao", ""),
        data.get("noi_dung_chi_dao", ""),
        data.get("chu_tri", ""),
        data.get("phoi_hop", ""),
        data.get("thu_muc_luu", ""),
        data.get("ten_file_luu", ""),
        data.get("thoi_gian_luu", ""),
    ]
    ws.append(row)

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[ws.max_row]:
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    saved_folder = data.get("thu_muc_luu", "")
    saved_name = data.get("ten_file_luu", "")
    if saved_folder and saved_name:
        file_cell = ws.cell(row=ws.max_row, column=12)
        saved_path = str(Path(saved_folder) / saved_name)
        file_cell.hyperlink = file_uri_from_path(saved_path)
        file_cell.style = "Hyperlink"
        file_cell.comment = None

    apply_excel_layout(ws)
    wb.save(LOG_FILE)
    wb.close()
    print(f"✅ Đã ghi Excel dòng STT {stt}: {data.get('so_vb', '')}")


def get_next_excel_stt() -> int:
    init_excel_log()
    wb = load_workbook(LOG_FILE, read_only=False)
    ws = wb[SHEET_NAME]
    stt = get_next_stt(ws)
    wb.close()
    return stt


# ==================================================
# HÀM PHỤ
# ==================================================


def wait(page, ms: int = 1000) -> None:
    page.wait_for_timeout(ms)


def clean_text(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n\s+", "\n", text)
    return text.strip()


def clean_date_vb(text: str) -> str:
    text = clean_text(text)
    m = re.search(r"\d{2}/\d{2}/\d{4}", text)
    return m.group(0) if m else text


def normalize_key(text: str) -> str:
    text = clean_text(text or "")
    text = text.lower()
    text = re.sub(r"\s+", "", text)
    return text


def build_duplicate_key(data: Dict[str, str]) -> str:
    so_vb = normalize_key(data.get("so_vb", ""))
    ngay_vb = normalize_key(data.get("ngay_vb", ""))
    if DUPLICATE_CHECK_MODE == "so_vb_time":
        tg = normalize_key(data.get("thoi_gian_chi_dao", ""))
        return f"{so_vb}|{tg}"
    if DUPLICATE_CHECK_MODE == "so_vb":
        return so_vb
    return f"{so_vb}|{ngay_vb}"


def load_existing_duplicate_keys() -> Set[str]:
    init_excel_log()
    keys: Set[str] = set()

    wb = load_workbook(LOG_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row or len(row) < 7:
            continue
        so_vb = str(row[1] or "")
        ngay_vb = str(row[2] or "")
        thoi_gian = str(row[6] or "")
        if not so_vb.strip():
            continue
        if DUPLICATE_CHECK_MODE == "so_vb_time":
            keys.add(f"{normalize_key(so_vb)}|{normalize_key(thoi_gian)}")
        elif DUPLICATE_CHECK_MODE == "so_vb":
            keys.add(normalize_key(so_vb))
        else:
            keys.add(f"{normalize_key(so_vb)}|{normalize_key(ngay_vb)}")

    wb.close()
    return keys


def load_existing_filenames() -> Set[str]:
    init_excel_log()
    names: Set[str] = set()
    wb = load_workbook(LOG_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        filename = str(row[11] or "") if row and len(row) > 11 else ""
        if filename.strip():
            names.add(normalize_key(filename))
    wb.close()
    return names


def parse_chi_dao_block(block_text: str) -> Dict[str, str]:
    """
    Dạng thường gặp:
    Trịnh Đình Chính - 17/05/2026 07:35:58
    Góp ý
    Chủ trì: TKT
    Phối hợp: P.TTĐ_TBA, P.TTĐ_DD
    """
    result = {
        "nguoi_chi_dao": "",
        "thoi_gian_chi_dao": "",
        "noi_dung_chi_dao": "",
        "chu_tri": "",
        "phoi_hop": "",
    }

    block_text = clean_text(block_text)
    if not block_text:
        return result

    lines = [clean_text(x) for x in block_text.splitlines() if clean_text(x)]
    if not lines:
        return result

    first = lines[0]
    m = re.match(r"^(.*?)\s*-\s*(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})$", first)
    if m:
        result["nguoi_chi_dao"] = clean_text(m.group(1))
        result["thoi_gian_chi_dao"] = clean_text(m.group(2))
    else:
        result["nguoi_chi_dao"] = first

    noi_dung_lines = []
    for line in lines[1:]:
        lower = line.lower()
        if lower.startswith("chủ trì:"):
            result["chu_tri"] = clean_text(line.split(":", 1)[1])
        elif lower.startswith("phối hợp:"):
            result["phoi_hop"] = clean_text(line.split(":", 1)[1])
        else:
            noi_dung_lines.append(line)

    result["noi_dung_chi_dao"] = "\n".join(noi_dung_lines)
    return result


def remove_vietnamese_accents(text: str) -> str:
    text = text.replace("Đ", "D").replace("đ", "d")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return unicodedata.normalize("NFC", text)


def safe_ascii_filename(name: str) -> str:
    name = remove_vietnamese_accents(name or "")
    name = name.replace("/", " ").replace("\\", " ")
    name = re.sub(r"[^A-Za-z0-9 ._\-]+", " ", name)
    name = re.sub(r"\s+", " ", name).strip(" ._-")
    return name or f"van_ban_{int(time.time())}"


def extract_so_vb_code(so_vb: str) -> str:
    so_vb = clean_text(so_vb or "")
    m = re.match(r"^\s*(\d+)", so_vb)
    if not m:
        m = re.search(r"\d+", so_vb)
    if not m:
        return "0000"
    digits = m.group(0)
    return digits.zfill(4) if len(digits) < 4 else digits


def add_date_prefix_if_needed(filename: str, so_vb: str) -> str:
    if re.match(r"^\d{6}-\d{4,}\s+-\s+", filename):
        return filename
    prefix = datetime.now().strftime("%y%m%d")
    vb_code = extract_so_vb_code(so_vb)
    return f"{prefix}-{vb_code} - {filename}"


def make_friendly_pdf_name(data: Dict[str, str], suggested_filename: str, stt: int) -> str:
    """Tạo tên PDF thân thiện: yymmdd-sốVB - tên gợi ý từ DOffice/PDF viewer."""
    base = suggested_filename or f"van_ban_{stt}.pdf"
    base = safe_ascii_filename(base)
    if not base.lower().endswith(".pdf"):
        base += ".pdf"
    return add_date_prefix_if_needed(base, data.get("so_vb", ""))


def unique_path(folder: Path, filename: str) -> Path:
    filename = safe_ascii_filename(filename)
    if not filename.lower().endswith(".pdf"):
        filename += ".pdf"

    path = folder / filename
    if not path.exists():
        return path

    stem = path.stem
    suffix = path.suffix
    for i in range(1, 9999):
        new_path = folder / f"{stem}_{i}{suffix}"
        if not new_path.exists():
            return new_path

    raise RuntimeError("Không tạo được tên file không trùng.")


def save_debug(page, name: str) -> None:
    path = f"debug_vb_phoi_hop_{name}.png"
    try:
        page.screenshot(path=path, full_page=True)
        print(f"📸 Đã lưu ảnh debug: {path}")
    except Exception as e:
        print(f"⚠️ Không chụp được ảnh debug: {e}")


def safe_click(locator, desc: str, timeout: int = 8000) -> bool:
    try:
        locator.wait_for(state="visible", timeout=timeout)
        locator.click(timeout=timeout)
        print(f"✅ Click: {desc}")
        return True
    except Exception as e:
        print(f"❌ Không click được {desc}: {e}")
        return False


# ==================================================
# ĐIỀU HƯỚNG DOffice
# ==================================================


def click_login_if_needed(page) -> None:
    """Nếu DOffice hiện nút Đăng nhập thì bấm; nếu đã vào sẵn thì bỏ qua."""
    try:
        btn = page.get_by_role("button", name=re.compile(r"Đăng nhập", re.I))
        if btn.count() > 0:
            btn.first.click(timeout=5000)
            print("✅ Đã click Đăng nhập.")
            wait(page, 2000)
    except Exception:
        pass


def choose_role_if_needed(page) -> None:
    print("\n--- Kiểm tra/chọn đúng chức danh ---")
    try:
        btn = page.get_by_role("button", name=re.compile(r"Nguyễn Xuân Bình.*Phòng ban", re.I))
        btn.wait_for(state="visible", timeout=8000)
        btn.click()
        wait(page, 500)

        role = page.get_by_role("menuitem", name=re.compile(r"Phó Truyền tải điện", re.I))
        role.wait_for(state="visible", timeout=5000)
        role.click()
        print("✅ Đã chọn chức danh Phó Truyền tải điện.")
        wait(page, 1500)
    except Exception as e:
        print("⚠️ Không chọn lại chức danh được hoặc không cần chọn:", e)


def open_phoi_hop_list(page) -> bool:
    print("\n--- Vào Văn bản / Chờ xử lý / Phối hợp ---")
    page.goto(DOFFICE_URL, wait_until="domcontentloaded")
    wait(page, 2500)

    click_login_if_needed(page)
    choose_role_if_needed(page)

    try:
        page.locator("fuse-vertical-navigation").get_by_text("Văn bản", exact=True).click(timeout=8000)
        print("✅ Click sidebar Văn bản trong fuse-vertical-navigation.")
    except Exception as e:
        print("⚠️ Không click được sidebar Văn bản bằng fuse navigation:", e)
        try:
            # Theo codegen có thể là nth(3), còn một số màn hình cũ là nth(1).
            page.locator("div").filter(has_text="Văn bản").nth(3).click(timeout=8000)
            print("✅ Click sidebar Văn bản bằng locator div filter nth(3).")
        except Exception as e2:
            print("⚠️ Không click được sidebar Văn bản bằng nth(3), thử nth(1):", e2)
            try:
                page.locator("div").filter(has_text="Văn bản").nth(1).click(timeout=8000)
                print("✅ Click sidebar Văn bản bằng locator div filter nth(1).")
            except Exception as e3:
                print("❌ Không click được sidebar Văn bản:", e3)
                save_debug(page, "sidebar_van_ban_failed")
                return False

    wait(page, 1000)

    if not safe_click(page.get_by_role("link", name=re.compile(r"Chờ xử lý", re.I)), "Link Chờ xử lý", timeout=10000):
        save_debug(page, "link_cho_xu_ly_failed")
        return False

    wait(page, 2000)

    try:
        page.get_by_role("tab", name=re.compile(r"Phối hợp", re.I)).click(timeout=10000)
        print("✅ Click tab Phối hợp.")
    except Exception as e:
        print("⚠️ Không click được tab Phối hợp bằng role, thử bằng text:", e)
        try:
            page.get_by_text(re.compile(r"Phối hợp\s*\(\d+\)", re.I)).click(timeout=10000)
            print("✅ Click tab Phối hợp bằng text.")
        except Exception as e2:
            print("⚠️ Không click được tab Phối hợp, kiểm tra xem danh sách có sẵn không:", e2)

    try:
        page.locator("tr.mat-row").first.wait_for(state="visible", timeout=12000)
        print("✅ Đã thấy danh sách văn bản Phối hợp.")
        wait(page, 1000)
        return True
    except Exception as e:
        print("❌ Không thấy danh sách văn bản:", e)
        save_debug(page, "phoi_hop_list_not_found")
        return False


# ==================================================
# TRÍCH XUẤT DỮ LIỆU ROW
# ==================================================


def scroll_document_list_down(page, pixels: int = 900) -> bool:
    moved = False
    try:
        box = page.locator("xly-vbleft, table, tr.mat-row").first.bounding_box(timeout=3000)
        if box:
            x = box["x"] + min(box["width"] / 2, 260)
            y = box["y"] + min(max(box["height"] / 2, 250), 650)
            page.mouse.move(x, y)
            page.mouse.wheel(0, pixels)
            moved = True
    except Exception:
        pass

    try:
        js_moved = page.evaluate(
            """
            (pixels) => {
                const els = Array.from(document.querySelectorAll('*'));
                let moved = false;
                for (const el of els) {
                    const r = el.getBoundingClientRect();
                    const style = window.getComputedStyle(el);
                    const canScroll = el.scrollHeight > el.clientHeight + 20;
                    const inLeftDocPanel = r.left >= 120 && r.left < 700 && r.top < window.innerHeight - 80 && r.bottom > 120;
                    const overflowY = style.overflowY;
                    if (canScroll && inLeftDocPanel && ['auto','scroll','overlay','hidden'].includes(overflowY)) {
                        const old = el.scrollTop;
                        el.scrollTop = old + pixels;
                        if (el.scrollTop !== old) moved = true;
                    }
                }
                return moved;
            }
            """,
            pixels,
        )
        moved = moved or bool(js_moved)
    except Exception:
        pass

    wait(page, 1200)
    return moved


def get_document_row(page, index_zero_based: int):
    rows = page.locator("tr.mat-row")
    for attempt in range(1, 13):
        count = rows.count()
        if count > index_zero_based:
            row = rows.nth(index_zero_based)
            try:
                row.scroll_into_view_if_needed(timeout=5000)
                row.wait_for(state="visible", timeout=5000)
                return row
            except Exception:
                pass

        print(f"⚠️ Chưa thấy row thứ {index_zero_based + 1}. DOM hiện có {count} row. Thử cuộn vùng danh sách lần {attempt}/12...")
        moved = scroll_document_list_down(page, pixels=900)
        if not moved:
            print("⚠️ Chưa xác định được container scroll; thử PageDown trên vùng danh sách...")
            try:
                page.keyboard.press("PageDown")
            except Exception:
                pass
            wait(page, 1000)

    raise RuntimeError(f"Không lấy được văn bản thứ {index_zero_based + 1} trong danh sách.")


def extract_document_info_from_row(row) -> Dict[str, str]:
    """Trích dữ liệu từ row văn bản chuyên môn / Văn bản / Chờ xử lý / Phối hợp."""
    data = {
        "so_vb": "",
        "ngay_vb": "",
        "noi_phat_hanh": "",
        "trich_yeu": "",
        "nguoi_chi_dao": "",
        "thoi_gian_chi_dao": "",
        "noi_dung_chi_dao": "",
        "chu_tri": "",
        "phoi_hop": "",
    }

    try:
        raw = row.evaluate(
            r"""
            (row) => {
                const clean = (s) => (s || '')
                    .replace(/\u00a0/g, ' ')
                    .replace(/[ \t]+/g, ' ')
                    .replace(/\n\s+/g, '\n')
                    .trim();

                const vb = row.querySelector('div.vb-item');
                if (!vb) return {};

                const directChildren = Array.from(vb.children);
                const directDivs = directChildren.filter(el => el.tagName && el.tagName.toLowerCase() === 'div');

                const firstDiv = directDivs[0] || null;
                const firstText = clean(firstDiv ? firstDiv.innerText : '');

                let soVb = '';
                if (firstDiv) {
                    const firstSpan = firstDiv.querySelector(':scope > span');
                    soVb = clean(firstSpan ? firstSpan.innerText : '');
                    if (!soVb && firstText) soVb = clean(firstText.split('\n')[0]);
                }

                let ngayVb = '';
                const dateMatch = firstText.match(/\d{2}\/\d{2}\/\d{4}/);
                if (dateMatch) ngayVb = dateMatch[0];

                const secondDiv = directDivs[1] || null;
                let noiPhatHanh = '';
                if (secondDiv) {
                    const blue = secondDiv.querySelector('span.text-blue-600, span[style*="color"]');
                    const firstSpan = secondDiv.querySelector(':scope > span');
                    noiPhatHanh = clean((blue || firstSpan || secondDiv).innerText);
                    noiPhatHanh = clean(noiPhatHanh.split('\n')[0]);
                }

                let trichYeu = '';
                const directSpan = directChildren.find(el => el.tagName && el.tagName.toLowerCase() === 'span');
                if (directSpan) trichYeu = clean(directSpan.innerText);
                if (!trichYeu) {
                    const unread = vb.querySelector('span.unread');
                    trichYeu = clean(unread ? unread.innerText : '');
                }
                if (!trichYeu) {
                    const thirdDiv = directDivs[2] || null;
                    if (thirdDiv) trichYeu = clean(thirdDiv.innerText.split('\n')[0]);
                }

                const chiDaoEl = vb.querySelector('section.text-blue-600 section')
                    || vb.querySelector('section.text-blue-600')
                    || vb.querySelector('lib-view-noi-dung-thuc-hien')
                    || null;
                const chiDaoText = clean(chiDaoEl ? chiDaoEl.innerText : '');

                return {
                    so_vb: soVb,
                    ngay_vb: ngayVb,
                    noi_phat_hanh: noiPhatHanh,
                    trich_yeu: trichYeu,
                    chi_dao_text: chiDaoText,
                    row_text: clean(vb.innerText),
                };
            }
            """
        ) or {}
    except Exception as e:
        print(f"⚠️ Không evaluate được row để trích metadata: {e}")
        raw = {}

    data["so_vb"] = clean_text(raw.get("so_vb", ""))
    data["ngay_vb"] = clean_date_vb(raw.get("ngay_vb", ""))
    data["noi_phat_hanh"] = clean_text(raw.get("noi_phat_hanh", ""))
    data["trich_yeu"] = clean_text(raw.get("trich_yeu", ""))
    data.update(parse_chi_dao_block(raw.get("chi_dao_text", "")))

    # Fallback cuối nếu DOM render khác.
    if not data["so_vb"] or not data["ngay_vb"] or not data["trich_yeu"]:
        row_text = clean_text(raw.get("row_text", ""))
        lines = [clean_text(x) for x in row_text.splitlines() if clean_text(x)]
        if not data["so_vb"] and lines:
            data["so_vb"] = lines[0]
        if not data["ngay_vb"]:
            m = re.search(r"\d{2}/\d{2}/\d{4}", row_text)
            if m:
                data["ngay_vb"] = m.group(0)
        if not data["noi_phat_hanh"]:
            for line in lines[1:5]:
                if re.fullmatch(r"\d{2}/\d{2}/\d{4}", line):
                    continue
                if line.lower().startswith(("chủ trì:", "phối hợp:")):
                    continue
                if re.search(r"\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2}", line):
                    continue
                if line != data["so_vb"]:
                    data["noi_phat_hanh"] = line
                    break
        if not data["trich_yeu"]:
            for line in lines:
                lower = line.lower()
                if line in (data["so_vb"], data["ngay_vb"], data["noi_phat_hanh"]):
                    continue
                if lower.startswith(("chủ trì:", "phối hợp:")):
                    continue
                if re.search(r"\d{2}/\d{2}/\d{4}", line):
                    continue
                data["trich_yeu"] = line
                break

    return data


def print_document_info(data: Dict[str, str]) -> None:
    print("\n📄 THÔNG TIN VĂN BẢN")
    print(f"  - Số VB:              {data.get('so_vb', '')}")
    print(f"  - Ngày VB:            {data.get('ngay_vb', '')}")
    print(f"  - Nơi phát hành:      {data.get('noi_phat_hanh', '')}")
    print(f"  - Trích yếu:          {data.get('trich_yeu', '')}")
    print(f"  - Người chỉ đạo:      {data.get('nguoi_chi_dao', '')}")
    print(f"  - Thời gian chỉ đạo:  {data.get('thoi_gian_chi_dao', '')}")
    print(f"  - Nội dung chỉ đạo:   {data.get('noi_dung_chi_dao', '')}")
    print(f"  - Chủ trì:            {data.get('chu_tri', '')}")
    print(f"  - Phối hợp:           {data.get('phoi_hop', '')}")


# ==================================================
# MỞ VĂN BẢN / TẢI PDF
# ==================================================


def click_document_row(row) -> bool:
    print("\n--- Mở văn bản đang xử lý ---")
    click_candidates = [
        (row.locator(".w-8").first, ".w-8 trong đúng row"),
        (row.locator("div.vb-item").first, "div.vb-item trong đúng row"),
        (row.locator("td.mat-cell").first, "td.mat-cell trong đúng row"),
    ]

    for locator, desc in click_candidates:
        try:
            locator.scroll_into_view_if_needed(timeout=3000)
            locator.click(timeout=8000)
            print(f"✅ Đã click văn bản bằng {desc}.")
            wait(row.page, 2500)
            return True
        except Exception as e:
            print(f"⚠️ Không click được {desc}: {e}")

    save_debug(row.page, "cannot_open_document_row")
    return False


def wait_pdf_ready(page) -> bool:
    print("--- Chờ PDF viewer sẵn sàng ---")
    try:
        download_btn = page.get_by_role("button", name=re.compile(r"Tải xuống|Download", re.I))
        download_btn.wait_for(state="visible", timeout=15000)
        wait(page, 800)
        print("✅ PDF viewer đã sẵn sàng, thấy nút Tải xuống.")
        return True
    except Exception as e:
        print("⚠️ Chưa thấy nút Tải xuống bằng role:", e)

    try:
        page.locator("#download").wait_for(state="visible", timeout=8000)
        wait(page, 800)
        print("✅ PDF viewer đã sẵn sàng, thấy #download.")
        return True
    except Exception as e:
        print("❌ Không thấy nút tải PDF:", e)
        save_debug(page, "pdf_not_ready")
        return False


def click_download_button(page) -> None:
    try:
        btn = page.get_by_role("button", name=re.compile(r"Tải xuống|Download", re.I))
        btn.wait_for(state="visible", timeout=5000)
        btn.click(timeout=5000)
        print("✅ Click nút Tải xuống bằng role.")
        return
    except Exception as e:
        print("⚠️ Role Tải xuống không click được:", e)

    try:
        btn = page.locator("#download").first
        btn.wait_for(state="visible", timeout=5000)
        btn.click(timeout=5000)
        print("✅ Click nút Tải xuống bằng #download.")
        return
    except Exception as e:
        print("⚠️ #download không click được:", e)

    page.evaluate(
        """
        () => {
            const btn = document.querySelector('#download')
                || document.querySelector('viewer-download-controls #download')
                || Array.from(document.querySelectorAll('button')).find(b => /Tải xuống|Download/i.test(b.innerText || b.getAttribute('aria-label') || ''));
            if (!btn) throw new Error('Không tìm thấy nút download trong DOM');
            btn.click();
        }
        """
    )
    print("✅ Click nút Tải xuống bằng JS fallback.")


def download_current_document(page, data: Dict[str, str], stt: int) -> Optional[Path]:
    if not ENABLE_DOWNLOAD_PDF:
        print("⏸️ ENABLE_DOWNLOAD_PDF = False, bỏ qua tải PDF.")
        return None

    print("\n--- Tải văn bản PDF ---")
    if not wait_pdf_ready(page):
        return None

    total_attempts = len(DOWNLOAD_ATTEMPT_TIMEOUTS_MS)
    for attempt, timeout_ms in enumerate(DOWNLOAD_ATTEMPT_TIMEOUTS_MS, start=1):
        print(f"Thử tải lần {attempt}/{total_attempts} | timeout {timeout_ms/1000:.0f}s...")
        try:
            with page.expect_download(timeout=timeout_ms) as download_info:
                click_download_button(page)

            download = download_info.value
            suggested = download.suggested_filename or f"van_ban_{int(time.time())}.pdf"
            friendly_name = make_friendly_pdf_name(data, suggested, stt)
            target = unique_path(DOWNLOAD_DIR, friendly_name)
            download.save_as(str(target))

            print(f"✅ Đã lưu file: {target}")
            return target

        except PlaywrightTimeoutError:
            print("⚠️ Không bắt được download event ở lần này, chuyển nhanh sang lần kế tiếp.")
            wait(page, 700)
        except Exception as e:
            print("⚠️ Lỗi tải file:", e)
            wait(page, 1000)

    print("❌ Tải file thất bại sau các lần thử.")
    save_debug(page, "download_failed")
    return None


# ==================================================
# KẾT THÚC VĂN BẢN
# ==================================================


def click_finish_quick_button(page) -> bool:
    candidates = []
    try:
        candidates.append((page.get_by_role("button", description=re.compile(r"K[eêế]t thúc nhanh", re.I)), "role button description Kết thúc nhanh"))
    except TypeError:
        pass

    candidates.extend([
        (page.get_by_role("button", name=re.compile(r"K[eêế]t thúc nhanh", re.I)), "role button name Kết thúc nhanh"),
        (page.locator(".fab-button-ktn > .mat-focus-indicator").first, "floating .fab-button-ktn"),
    ])

    for locator, desc in candidates:
        try:
            locator.wait_for(state="visible", timeout=8000)
            locator.click(timeout=8000)
            print(f"✅ Đã click nút Kết thúc văn bản bằng {desc}.")
            return True
        except Exception as e:
            print(f"⚠️ Không click được {desc}: {e}")

    save_debug(page, "finish_button_failed")
    return False


def click_save_finish_button(page) -> bool:
    wait(page, 3500)

    dialog = page.locator("mat-dialog-container, .mat-dialog-container, [role='dialog']").last
    candidates = [
        (dialog.get_by_role("button", name=re.compile(r"Lưu", re.I)), "nút Lưu trong dialog"),
        (page.get_by_role("button", name=re.compile(r"Lưu", re.I)), "role button có chữ Lưu"),
        (page.locator("button").filter(has_text=re.compile(r"Lưu", re.I)).last, "button filter has_text Lưu"),
    ]

    for locator, desc in candidates:
        try:
            locator.wait_for(state="visible", timeout=12000)
            locator.click(timeout=12000)
            print(f"✅ Đã click Lưu xác nhận kết thúc bằng {desc}.")
            return True
        except Exception as e:
            print(f"⚠️ Không click được {desc}: {e}")

    try:
        page.evaluate(
            """
            () => {
                const buttons = Array.from(document.querySelectorAll('button'));
                const btn = buttons.reverse().find(b => /Lưu/i.test((b.innerText || b.textContent || '').trim()));
                if (!btn) throw new Error('Không tìm thấy button Lưu trong DOM');
                btn.click();
            }
            """
        )
        print("✅ Đã click Lưu xác nhận kết thúc bằng JS fallback.")
        return True
    except Exception as e:
        print("❌ Không click được nút Lưu:", e)
        save_debug(page, "save_finish_failed")
        return False


def finish_current_document(page) -> bool:
    print("\n--- Kết thúc văn bản ---")

    if not ENABLE_FINISH_DOCUMENT:
        print("⏸️ ENABLE_FINISH_DOCUMENT = False, bỏ qua kết thúc văn bản.")
        return True

    if ASK_CONFIRM_BEFORE_FINISH:
        ans = input("Chuẩn bị bấm KẾT THÚC văn bản này. Đồng ý? y/n: ").strip().lower()
        if ans != "y":
            print("⏸️ Người dùng bỏ qua kết thúc văn bản.")
            return False

    if not click_finish_quick_button(page):
        return False

    try:
        page.get_by_text(re.compile(r"Kết thúc nhanh", re.I)).wait_for(state="visible", timeout=12000)
        print("✅ Đã thấy popup Kết thúc nhanh.")
    except Exception:
        print("⚠️ Chưa bắt được text Kết thúc nhanh, vẫn tiếp tục chờ nút Lưu.")

    if ASK_CONFIRM_BEFORE_FINISH:
        ans = input("Chuẩn bị bấm LƯU để xác nhận kết thúc. Đồng ý? y/n: ").strip().lower()
        if ans != "y":
            print("⏸️ Người dùng không xác nhận Lưu.")
            return False

    if not click_save_finish_button(page):
        return False

    try:
        page.locator("mat-dialog-container, .mat-dialog-container, [role='dialog']").last.wait_for(state="hidden", timeout=15000)
    except Exception:
        pass
    wait(page, 3000)
    return True


# ==================================================
# MAIN
# ==================================================


def main() -> None:
    if not AUTH_STATE.exists():
        raise FileNotFoundError(
            f"Không thấy file session: {AUTH_STATE}\n"
            "Chạy login_save_state.py trước."
        )

    init_excel_log()
    existing_keys = load_existing_duplicate_keys()
    existing_filenames = load_existing_filenames()

    print("DOWNLOAD_DIR:", DOWNLOAD_DIR)
    print("LOG_FILE:", LOG_FILE)
    print("MAX_DOCUMENTS:", MAX_DOCUMENTS)
    print("ENABLE_DOWNLOAD_PDF:", ENABLE_DOWNLOAD_PDF)
    print("STOP_WHEN_DUPLICATE_FOUND:", STOP_WHEN_DUPLICATE_FOUND)
    print("DUPLICATE_CHECK_MODE:", DUPLICATE_CHECK_MODE)
    print("ENABLE_FINISH_DOCUMENT:", ENABLE_FINISH_DOCUMENT)
    print("ASK_CONFIRM_BEFORE_FINISH:", ASK_CONFIRM_BEFORE_FINISH)
    print(f"📘 Excel hiện có {len(existing_keys)} khóa văn bản đã tổng hợp.")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=SLOW_MO_MS)
        context = browser.new_context(
            storage_state=str(AUTH_STATE),
            accept_downloads=True,
            viewport={"width": 1600, "height": 900},
        )
        page = context.new_page()

        try:
            if not open_phoi_hop_list(page):
                print("Không vào được danh sách Phối hợp.")
                return

            processed = 0
            current_index = 0

            while processed < MAX_DOCUMENTS:
                print("\n" + "=" * 70)
                print(f"XỬ LÝ VĂN BẢN THỨ {processed + 1} | ROW ĐANG CHỌN {current_index + 1} | ĐÃ GHI {processed}/{MAX_DOCUMENTS}")
                print("=" * 70)

                if REFRESH_LIST_EVERY and processed > 0 and processed % REFRESH_LIST_EVERY == 0:
                    print(f"\n🔄 Đã xử lý {processed} văn bản. Load lại danh sách Phối hợp...")
                    if not open_phoi_hop_list(page):
                        print("Không load lại được danh sách Phối hợp. Dừng.")
                        break
                    current_index = 0
                    wait(page, 2500)

                try:
                    row_index_to_pick = 0 if (ENABLE_FINISH_DOCUMENT and ALWAYS_PICK_FIRST_ROW_AFTER_FINISH) else current_index
                    row = get_document_row(page, row_index_to_pick)
                except Exception as e:
                    print("❌ Không lấy được row văn bản:", e)
                    print("🔄 Thử load lại danh sách Phối hợp một lần nữa...")
                    if not open_phoi_hop_list(page):
                        save_debug(page, "cannot_reload_list")
                        break
                    wait(page, 2500)
                    try:
                        row_index_to_pick = 0 if (ENABLE_FINISH_DOCUMENT and ALWAYS_PICK_FIRST_ROW_AFTER_FINISH) else current_index
                        row = get_document_row(page, row_index_to_pick)
                    except Exception as e2:
                        print("❌ Vẫn không lấy được row văn bản sau khi reload:", e2)
                        save_debug(page, "cannot_get_document_row_after_reload")
                        break

                if not click_document_row(row):
                    print("Không mở được văn bản. Dừng để kiểm tra.")
                    break

                data = extract_document_info_from_row(row)
                print_document_info(data)

                duplicate_key = build_duplicate_key(data)
                if duplicate_key and duplicate_key in existing_keys:
                    print(f"🛑 Văn bản đã có trong Excel: {data.get('so_vb', '')}")
                    if STOP_WHEN_DUPLICATE_FOUND:
                        print("Dừng chương trình vì đã gặp văn bản trùng. Danh sách đang sắp xếp mới nhất ở trên nên các dòng dưới thường đã tổng hợp rồi.")
                        break
                    current_index += 1
                    continue

                planned_stt = get_next_excel_stt()
                saved_file = download_current_document(page, data, planned_stt)
                if ENABLE_DOWNLOAD_PDF and not saved_file:
                    print("Không tải được văn bản. Dừng để kiểm tra.")
                    break

                if not finish_current_document(page):
                    print("Không kết thúc được văn bản. Dừng để kiểm tra.")
                    break

                if saved_file:
                    filename_key = normalize_key(saved_file.name)
                    if filename_key in existing_filenames:
                        print(f"🛑 Tên file đã có trong Excel: {saved_file.name}")
                        if STOP_WHEN_DUPLICATE_FOUND:
                            print("Dừng chương trình vì gặp tên file trùng.")
                            break
                    data["thu_muc_luu"] = str(saved_file.parent)
                    data["ten_file_luu"] = saved_file.name
                else:
                    data["thu_muc_luu"] = str(DOWNLOAD_DIR)
                    data["ten_file_luu"] = ""

                data["thoi_gian_luu"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                append_excel_log(data)

                existing_keys.add(duplicate_key)
                if data.get("ten_file_luu"):
                    existing_filenames.add(normalize_key(data["ten_file_luu"]))

                processed += 1
                if ENABLE_FINISH_DOCUMENT and ALWAYS_PICK_FIRST_ROW_AFTER_FINISH:
                    current_index = 0
                else:
                    current_index += 1
                wait(page, 1500)

            print(f"\n✅ Hoàn tất. Đã ghi mới {processed} văn bản.")
            print(f"📘 File tổng hợp: {LOG_FILE}")

        finally:
            if PAUSE_BEFORE_CLOSE:
                input("Nhấn Enter để đóng browser...")
            browser.close()


if __name__ == "__main__":
    main()
