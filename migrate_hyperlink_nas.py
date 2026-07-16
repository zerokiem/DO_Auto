"""Chuyen TOAN BO hyperlink cot "Ten file luu" trong Tong_hop_DOffice.xlsx sang
link web toi NAS (DISPLAY_BASE_URL, vd http://100.100.1.254:8877/vb) de bam MO
DUOC tren dien thoai co Tailscale va moi thiet bi trong mang.

Link moi = DISPLAY_BASE_URL/<thu_muc_con>/<ten_file>, trong do <thu_muc_con> lay
tu thanh phan cuoi cot "Thu muc luu" (vd VB_Chu_tri_da_XL) - dung duoc cho ca dong
cu (duong dan D:\\OneDrive...) lan dong moi (S:\\... hay /data/...).

Chi doi HYPERLINK; gia tri chu (cot "Thu muc luu", ten file) giu nguyen.

Chay TRONG container (da co env DOFFICE_DISPLAY_URL):
    sudo -n /usr/local/bin/docker exec doffice python /app/migrate_hyperlink_nas.py
"""
from __future__ import annotations

import shutil
from datetime import datetime

from openpyxl import load_workbook

import config
from do_auto import excel_log

EXCEL = config.EXCEL_FILE
LINK_BASE = str(getattr(config, "DISPLAY_BASE_URL", "") or "")

FOLDER_COL = 11  # "Thu muc luu"
FILE_COL = 12    # "Ten file luu" (chua hyperlink)


def main() -> None:
    print(f"EXCEL     = {EXCEL}")
    print(f"LINK_BASE = {LINK_BASE!r}")
    if not (LINK_BASE.startswith("http://") or LINK_BASE.startswith("https://")):
        print("!! DISPLAY_BASE_URL chua duoc dat (env DOFFICE_DISPLAY_URL). Can "
              "'docker-compose up -d' de nap env moi truoc khi chay. Dung.")
        return
    if not EXCEL.exists():
        print("Khong thay file Excel:", EXCEL)
        return

    backup = EXCEL.with_name(EXCEL.stem + f".backup_hlink_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(EXCEL, backup)
    print("Da sao luu:", backup)

    wb = load_workbook(EXCEL)
    total = 0
    for ws in wb.worksheets:
        fixed_here = 0
        for r in range(excel_log.DATA_START_ROW, ws.max_row + 1):
            folder = str(ws.cell(r, FOLDER_COL).value or "")
            file_cell = ws.cell(r, FILE_COL)
            fname = str(file_cell.value or "")
            if not fname.strip():
                continue  # dong khong co file (tai loi) -> khong tao link
            file_cell.hyperlink = excel_log.build_file_link(folder, fname, LINK_BASE)
            file_cell.style = "Hyperlink"
            fixed_here += 1
        total += fixed_here
        print(f"  [{ws.title}] doi {fixed_here} hyperlink")

    wb.save(EXCEL)
    wb.close()
    print(f"XONG. Da doi tong cong {total} hyperlink trong {EXCEL}")


if __name__ == "__main__":
    main()
