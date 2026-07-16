"""Sua cac dong Excel CU dang luu duong dan noi bo cua container (/data/...) thanh
duong dan HIEN THI (S:\\Working\\Van_ban\\... theo DOFFICE_DISPLAY_DIR) de nguoi
dung bam MO DUOC file tren may Windows.

Chi sua nhung dong co "Thu muc luu" nam DUOI DOWNLOAD_BASE_DIR hien tai (tuc
/data khi chay trong container). Cac dong duong dan Windows cu (vd
D:\\OneDrive - NPT\\9. Jobs\\Van_ban\\...) duoc GIU NGUYEN, vi file do nam trong
OneDrive tren may nguoi dung chu khong nam tren NAS.

Chay TRONG container (da co openpyxl + env DOFFICE_DISPLAY_DIR):
    sudo -n /usr/local/bin/docker exec doffice python /app/migrate_display_path.py
"""
from __future__ import annotations

import shutil
from datetime import datetime

from openpyxl import load_workbook

import config
from do_auto import excel_log

EXCEL = config.EXCEL_FILE
DOWNLOAD_BASE = str(config.DOWNLOAD_BASE_DIR)  # /data trong container
DISPLAY_BASE = str(config.DISPLAY_BASE_DIR)    # S:\Working\Van_ban

FOLDER_COL = 11  # cot "Thu muc luu"
FILE_COL = 12    # cot "Ten file luu" (chua hyperlink)


def needs_fix(folder: str) -> bool:
    f = (folder or "").replace("\\", "/").rstrip("/")
    b = DOWNLOAD_BASE.replace("\\", "/").rstrip("/")
    return bool(f) and (f == b or f.startswith(b + "/"))


def main() -> None:
    print(f"EXCEL        = {EXCEL}")
    print(f"DOWNLOAD_BASE= {DOWNLOAD_BASE}")
    print(f"DISPLAY_BASE = {DISPLAY_BASE}")
    if DOWNLOAD_BASE.replace('\\', '/').rstrip('/') == DISPLAY_BASE.replace('\\', '/').rstrip('/'):
        print("!! DISPLAY_BASE trung DOWNLOAD_BASE -> khong co gi de doi. "
              "Kiem tra bien moi truong DOFFICE_DISPLAY_DIR da duoc nap chua "
              "(can 'docker-compose up -d' de tao lai container voi env moi).")
        return
    if not EXCEL.exists():
        print("Khong thay file Excel:", EXCEL)
        return

    backup = EXCEL.with_name(EXCEL.stem + f".backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(EXCEL, backup)
    print("Da sao luu:", backup)

    wb = load_workbook(EXCEL)
    total = 0
    for ws in wb.worksheets:
        fixed_here = 0
        for r in range(excel_log.DATA_START_ROW, ws.max_row + 1):
            folder_cell = ws.cell(row=r, column=FOLDER_COL)
            folder = str(folder_cell.value or "")
            if not needs_fix(folder):
                continue
            new_folder = excel_log.to_display_folder(folder, DOWNLOAD_BASE, DISPLAY_BASE)
            folder_cell.value = new_folder

            file_cell = ws.cell(row=r, column=FILE_COL)
            fname = str(file_cell.value or "")
            if fname:
                sep = "\\" if ("\\" in new_folder or (len(new_folder) >= 2 and new_folder[1] == ":")) else "/"
                path = new_folder.rstrip("/\\") + sep + fname
                file_cell.hyperlink = excel_log.file_uri_from_path(path)
                file_cell.style = "Hyperlink"
            fixed_here += 1
        total += fixed_here
        print(f"  [{ws.title}] sua {fixed_here} dong")

    wb.save(EXCEL)
    wb.close()
    print(f"XONG. Da sua tong cong {total} dong trong {EXCEL}")


if __name__ == "__main__":
    main()
