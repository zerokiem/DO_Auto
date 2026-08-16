"""
Flask app cho DOffice web dashboard.

Day la cong cu NOI BO chay tren may cua ban (hoac 1 may trong mang noi bo/
Tailscale), KHONG co lop dang nhap rieng - vi Playwright/Chromium can chay
ngay tren may nay va DOffice chi truy cap duoc tu mang noi bo. Xem run_web.py
o thu muc goc de biet cach khoi dong.
"""
from __future__ import annotations

import json
import os
import queue
import sys
import time
from datetime import datetime
from pathlib import Path

from flask import Flask, Response, abort, jsonify, redirect, render_template, request, send_file, url_for

import config as base_config
from do_auto import excel_log, history, scheduler as scheduler_mod, settings_store
from webapp.login_manager import LoginManager
from webapp.run_manager import RunManager

PROJECT_DIR = Path(__file__).resolve().parent.parent

app = Flask(__name__)
app.jinja_env.globals["zip"] = zip
run_manager = RunManager(base_config)
login_manager = LoginManager(base_config)

# Cac route da duyet truc tiep tren DOffice v2.4.0. Input tren web van cho go
# tu do de khong khoa cung khi DOffice bo sung route moi.
DOFFICE_ROUTE_CHOICES = [
    ("Văn bản xử lý — Chờ xử lý", "/congviec/ld-xly-vb/ChoXL"),
    ("Văn bản xử lý — Đã xử lý", "/congviec/ld-xly-vb/DaXL"),
    ("Văn bản xử lý — Xem để biết", "/congviec/ld-xly-vb/XemDB"),
    ("Văn bản xử lý — Chuyển nhầm", "/congviec/ld-xly-vb/ChuyenNham"),
    ("Văn bản xử lý — Theo dõi", "/congviec/ld-xly-vb/TheoDoi"),
    ("Công việc ban/phòng — Chờ giao việc", "/congviec/cviec-ldphong-xu-ly/chogiao"),
    ("Công việc ban/phòng — Đã giao việc", "/congviec/cviec-ldphong-xu-ly/dagiao"),
    ("Công việc ban/phòng — Xem để biết", "/congviec/cviec-ldphong-xu-ly/xemdb"),
    ("Công việc ban/phòng — Danh sách chuyển nhầm", "/congviec/cviec-ldphong-xu-ly/chuyennham"),
    ("Công việc ban/phòng — Duyệt hạn", "/congviec/cviec-ldphong-xu-ly/duyethan"),
    ("Công việc ban/phòng — Theo dõi", "/congviec/cviec-ldphong-xu-ly/theodoi"),
    ("Công việc cá nhân — Chờ thực hiện", "/congviec/cviec-cvien-xuly/chothuchien"),
    ("Công việc cá nhân — Đang thực hiện", "/congviec/cviec-cvien-xuly/dangthuchien"),
    ("Công việc cá nhân — Đã thực hiện", "/congviec/cviec-cvien-xuly/dathuchien"),
    ("Công việc cá nhân — Xem để biết", "/congviec/cviec-cvien-xuly/xemdb"),
    ("Công việc cá nhân — Theo dõi", "/congviec/cviec-cvien-xuly/theodoi"),
    ("Văn bản đi — Chờ ký số", "/duthaovanban/danhsach/vbdi/choduyet"),
    ("Văn bản đi — Đã duyệt", "/duthaovanban/danhsach/vbdi/daduyet"),
    ("Văn bản đi — Đã phát hành", "/duthaovanban/danhsach/vbdi/phathanh"),
    ("Văn bản đi — Trả lại", "/duthaovanban/danhsach/vbdi/tralai"),
    ("Văn bản nội bộ — Chờ ký số", "/duthaovanban/danhsach/vbnb/choduyet"),
    ("Văn bản nội bộ — Đã duyệt", "/duthaovanban/danhsach/vbnb/daduyet"),
    ("Văn bản nội bộ — Đã phát hành", "/duthaovanban/danhsach/vbnb/phathanh"),
    ("Văn bản nội bộ — Trả lại", "/duthaovanban/danhsach/vbnb/tralai"),
    ("Hồ sơ — Danh mục hồ sơ", "/hstl/danh-muc-hso-nam"),
    ("Hồ sơ — Công việc cá nhân", "/hstl/hso-cong-viec"),
    ("Hồ sơ — Công việc phòng ban", "/hstl/hso-phong-ban"),
    ("Hồ sơ — Lưu trữ cơ quan", "/hstl/hso-co-quan"),
    ("Hồ sơ — Khai thác", "/hstl/khai-thac-hso"),
    ("Hồ sơ — Tìm kiếm", "/hstl/tim-kiem-hso"),
]


@app.context_processor
def inject_globals():
    return {"current_year": datetime.now().year}


def _ordered_tasks(cfg):
    return sorted(cfg.TASKS.items(), key=lambda kv: kv[1].sheet_order)


def _validate_navigation_fields(label, advanced, steps, direct_href, sidebar, list_link, tab_name, extract_mode):
    if advanced:
        if not steps:
            raise ValueError(f"Tác vụ '{label}' đã bật điều hướng nâng cao nhưng chưa có bước JSON.")
        return
    if not direct_href and (not sidebar or not list_link):
        raise ValueError(f"Tác vụ '{label}' cần chọn Đường dẫn trực tiếp hoặc nhập cả Sidebar và Tiểu mục.")
    if extract_mode == "directive" and not tab_name:
        raise ValueError(
            f"Tác vụ '{label}' cần nhập Tab Văn bản (ví dụ Chủ trì, Phối hợp, Chưa xử lý hoặc Đã xử lý)."
        )


def _all_excel_sheet_names(cfg) -> list[str]:
    """Tra ve sheet cua task dang bat ke ca sheet cu cua task da xoa.

    Xoa task chi xoa cau hinh chay, khong xoa lich su Excel. Hien cac sheet cu
    tren web de nguoi dung van tra cuu duoc va co bang chung rang du lieu khong
    bi mat.
    """
    names = [task.sheet_name for _, task in _ordered_tasks(cfg)]
    excel_file = Path(cfg.EXCEL_FILE)
    if not excel_file.exists():
        return names
    try:
        from openpyxl import load_workbook

        wb = load_workbook(excel_file, read_only=True)
        for name in wb.sheetnames:
            if name not in names:
                names.append(name)
        wb.close()
    except Exception:
        pass
    return names


def _parse_navigation_steps(raw: str, task_name: str) -> list[dict]:
    """Kiem tra JSON truoc khi ghi vao config.py, bao loi de hieu tren giao dien."""
    try:
        steps = json.loads(raw or "[]")
    except json.JSONDecodeError as e:
        raise ValueError(f"Chuỗi bước của '{task_name}' không phải JSON hợp lệ: {e.msg} (dòng {e.lineno}).") from e
    if not isinstance(steps, list) or any(not isinstance(step, dict) for step in steps):
        raise ValueError(f"Chuỗi bước của '{task_name}' phải là mảng JSON gồm các object.")
    return steps


def _can_login_interactively() -> bool:
    """True neu may nay co man hinh that de mo cua so Chromium (Windows, hoac
    Linux co bien DISPLAY) - quyet dinh dung luong dang nhap tuong tac hay
    headless (form username/password). Dung chung cho settings_page() va
    api_login_start() de khong bao gio lech logic giua 2 noi."""
    return (sys.platform == "win32") or bool(os.environ.get("DISPLAY"))


def _auth_state_info(cfg) -> dict:
    p = Path(cfg.AUTH_STATE)
    if not p.exists():
        return {"exists": False, "age_days": None}
    age_days = (time.time() - p.stat().st_mtime) / 86400
    return {"exists": True, "age_days": round(age_days, 1)}


def _read_sheet_preview(excel_file: Path, sheet_name: str, limit=300):
    """Doc sheet, tra ve (headers, rows_moi_nhat_len_tren, tong_so_dong).
    limit=None nghia la lay tat ca."""
    from openpyxl import load_workbook

    excel_file = Path(excel_file)
    if not excel_file.exists():
        return excel_log.HEADERS, [], 0

    wb = load_workbook(excel_file, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return excel_log.HEADERS, [], 0
    ws = wb[sheet_name]

    all_rows = []
    for row in ws.iter_rows(min_row=excel_log.DATA_START_ROW, values_only=True):
        if not row or not any(v not in (None, "") for v in row):
            continue
        all_rows.append(row)
    wb.close()

    total = len(all_rows)
    tail = all_rows if limit is None else all_rows[-limit:]
    tail = list(tail)
    tail.reverse()  # moi nhat len tren
    return excel_log.HEADERS, tail, total


@app.route("/")
def dashboard():
    effective_cfg = settings_store.build_effective_config(base_config)
    tasks = _ordered_tasks(effective_cfg)

    recent_by_task = {}
    for key, _task in tasks:
        rows = history.list_recent_runs_for_task(Path(effective_cfg.HISTORY_DB), key, limit=1)
        recent_by_task[key] = rows[0] if rows else None

    return render_template(
        "dashboard.html",
        tasks=tasks,
        recent_by_task=recent_by_task,
        auth=_auth_state_info(effective_cfg),
        status=run_manager.status(),
        login_status=login_manager.status(),
        excel_file=str(effective_cfg.EXCEL_FILE),
        is_windows=sys.platform == "win32",
    )


@app.post("/api/run")
def api_run():
    payload = request.get_json(force=True, silent=True) or {}
    task_keys = payload.get("task_keys") or []
    # Tren Linux/Docker (NAS) khong co man hinh de hien Chromium - ep headless=True
    # bat ke gia tri tich chon tren web, tranh loi/crash khi mo trinh duyet
    # "khong an" trong 1 container khong co display server.
    headless = True if sys.platform != "win32" else bool(payload.get("headless", True))
    test_mode = bool(payload.get("test_mode", False))

    effective_cfg = settings_store.build_effective_config(base_config)

    unknown = [k for k in task_keys if k not in effective_cfg.TASKS]
    if unknown:
        return jsonify({"started": False, "error": f"Tác vụ không hợp lệ: {', '.join(unknown)}"}), 400
    if not task_keys:
        return jsonify({"started": False, "error": "Chưa chọn tác vụ nào."}), 400
    if not Path(effective_cfg.AUTH_STATE).exists():
        return (
            jsonify(
                {
                    "started": False,
                    "error": "Chưa có phiên đăng nhập. Bấm 'Đăng nhập lại' ở trên hoặc chạy "
                    "'python login_save_state.py' trên máy này trước.",
                }
            ),
            400,
        )

    started = run_manager.start(task_keys, headless=headless, test_mode=test_mode)
    if not started:
        return jsonify({"started": False, "error": "Đang có 1 lượt chạy khác, vui lòng đợi."}), 409
    return jsonify({"started": True})


@app.get("/api/status")
def api_status():
    return jsonify(run_manager.status())


@app.get("/api/logs/stream")
def api_logs_stream():
    def gen():
        q = run_manager.broadcaster.subscribe()
        try:
            for line in list(run_manager.broadcaster.history)[-200:]:
                yield f"data: {json.dumps(line)}\n\n"
            while True:
                try:
                    line = q.get(timeout=15)
                    yield f"data: {json.dumps(line)}\n\n"
                except queue.Empty:
                    yield ": ping\n\n"
        except GeneratorExit:
            pass
        finally:
            run_manager.broadcaster.unsubscribe(q)

    return Response(gen(), mimetype="text/event-stream")


@app.post("/api/login/start")
def api_login_start():
    username = password = None
    if not _can_login_interactively():
        # May khong man hinh (NAS/Docker/Pi): BAT BUOC phai co username/password
        # tu form web de dang nhap headless - khong the mo cua so that.
        payload = request.get_json(force=True, silent=True) or {}
        username = (payload.get("username") or "").strip()
        password = payload.get("password") or ""
        if not username or not password:
            return jsonify({"started": False, "error": "Thiếu tài khoản hoặc mật khẩu."}), 400

    started = login_manager.start(username=username, password=password)
    if not started:
        return jsonify({"started": False, "error": "Đang có 1 lượt đăng nhập khác đang mở."}), 409
    return jsonify({"started": True})


@app.post("/api/login/confirm")
def api_login_confirm():
    ok = login_manager.confirm()
    if not ok:
        return jsonify({"ok": False, "error": "Chưa có lượt đăng nhập nào đang chờ xác nhận."}), 400
    return jsonify({"ok": True})


@app.get("/api/login/status")
def api_login_status():
    return jsonify(login_manager.status())


@app.post("/api/telegram/test")
def api_telegram_test():
    payload = request.get_json(force=True, silent=True) or {}
    bot_token = (payload.get("bot_token") or "").strip()
    chat_id = (payload.get("chat_id") or "").strip()

    from do_auto import notify

    test_message = "✅ DOffice - đây là tin nhắn thử. Nếu bạn thấy tin này, cấu hình Telegram đã đúng."
    ok, error = notify.send_telegram_message(bot_token, chat_id, test_message)
    if ok:
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": error}), 400


@app.route("/history")
def history_page():
    effective_cfg = settings_store.build_effective_config(base_config)
    rows = history.list_recent_runs(Path(effective_cfg.HISTORY_DB), limit=150)
    return render_template("history.html", rows=rows)


@app.get("/logs/<path:filename>")
def view_log(filename):
    effective_cfg = settings_store.build_effective_config(base_config)
    logs_dir = Path(effective_cfg.LOGS_DIR).resolve()
    target = (logs_dir / filename).resolve()
    # Chan doc file ngoai LOGS_DIR (vd truyen "../../config.py" qua filename).
    if logs_dir not in target.parents and target != logs_dir:
        abort(403)
    if not target.exists() or not target.is_file():
        abort(404)
    content = target.read_text(encoding="utf-8", errors="replace")
    return render_template("log_view.html", filename=filename, content=content)


@app.route("/excel")
def excel_page():
    effective_cfg = settings_store.build_effective_config(base_config)
    sheet_names = _all_excel_sheet_names(effective_cfg)

    sheet = request.args.get("sheet")
    if not sheet or sheet not in sheet_names:
        sheet = sheet_names[0] if sheet_names else None

    # So dong hien thi: tham so ?limit= (so nguyen hoac "all"). Mac dinh 50.
    limit_options = [20, 50, 100, 200, 300, 500]
    limit_arg = (request.args.get("limit") or "50").strip().lower()
    if limit_arg in ("all", "0", "tatca", "tat_ca"):
        limit_arg = "all"
        limit = None
    else:
        try:
            limit = max(1, int(limit_arg))
            limit_arg = str(limit)
        except ValueError:
            limit, limit_arg = 50, "50"

    columns, rows, error, total = [], [], None, 0
    file_links = []
    if sheet:
        try:
            columns, rows, total = _read_sheet_preview(effective_cfg.EXCEL_FILE, sheet, limit=limit)
        except Exception as e:
            error = str(e)

    # Xay LAI link cho cot "Ten file luu" ngay tu Thu muc luu + Ten file luu +
    # DISPLAY_BASE_URL hien tai (khong doc hyperlink da luu san trong file - che
    # do read_only cua openpyxl khong ho tro doc .hyperlink), de link tren web
    # luon dung theo cau hinh moi nhat kem ca voi du lieu vua sua qua migration.
    if columns and "Thư mục lưu" in columns and "Tên file lưu" in columns:
        folder_idx = columns.index("Thư mục lưu")
        file_idx = columns.index("Tên file lưu")
        for row in rows:
            fname = str(row[file_idx] or "").strip()
            if not fname:
                file_links.append(None)
                continue
            folder = str(row[folder_idx] or "")
            file_links.append(excel_log.build_file_link(folder, fname, effective_cfg.DISPLAY_BASE_URL))

    return render_template(
        "excel_view.html",
        sheet_names=sheet_names,
        active_sheet=sheet,
        columns=columns,
        rows=rows,
        file_links=file_links,
        error=error,
        row_count=len(rows),
        total_count=total,
        limit_options=limit_options,
        current_limit=limit_arg,
        column_widths=excel_log.COLUMN_WIDTHS,
    )


@app.get("/vb/<path:relpath>")
def serve_vb(relpath):
    """Phuc vu file PDF trong thu muc du lieu (DOWNLOAD_BASE_DIR = /data) qua web,
    de hyperlink trong Excel mo duoc tren dien thoai/Tailscale. Chi cho phep file
    NAM TRONG thu muc du lieu (chan '../' thoat ra ngoai)."""
    effective_cfg = settings_store.build_effective_config(base_config)
    base = Path(effective_cfg.DOWNLOAD_BASE_DIR).resolve()
    target = (base / relpath).resolve()
    if base != target and base not in target.parents:
        abort(403)
    if not target.exists() or not target.is_file():
        abort(404)
    return send_file(target)


@app.get("/excel/download")
def excel_download():
    effective_cfg = settings_store.build_effective_config(base_config)
    path = Path(effective_cfg.EXCEL_FILE)
    if not path.exists():
        return "Chưa có file Excel nào được tạo.", 404
    return send_file(path, as_attachment=True, download_name=path.name)


@app.route("/settings", methods=["GET", "POST"])
def settings_page():
    # Dung effective_cfg (qua settings_store, co getattr + fallback mac dinh)
    # thay vi doc truc tiep base_config.X lam gia tri fallback ben duoi - de
    # KHONG BAO GIO crash du config.py cua nguoi dung cu hon code hien tai.
    effective_cfg = settings_store.build_effective_config(base_config)

    if request.method == "POST":
        form_action = request.form.get("form_action", "save")
        try:
            if form_action.startswith("delete:"):
                task_key = form_action.split(":", 1)[1]
                settings_store.remove_task(task_key)
                settings_store.reload_config(base_config)
                # Chu y: khong goi bat ky ham xoa workbook/sheet o day.
                return redirect(url_for("settings_page", saved=1, message="Đã xóa tác vụ khỏi danh sách chạy."))

            if form_action == "create_task":
                new_label = request.form.get("new_label", "").strip()
                new_key = request.form.get("new_key", "").strip()
                new_steps = _parse_navigation_steps(request.form.get("new_navigation_steps", "[]"), new_label or new_key)
                new_advanced = request.form.get("new_use_advanced_navigation") == "on"
                new_sidebar = request.form.get("new_sidebar_item", "").strip()
                new_list_link = request.form.get("new_list_link", "").strip()
                new_direct_href = request.form.get("new_list_link_href", "").strip()
                new_tab_name = request.form.get("new_tab_name", "").strip()
                new_extract_mode = request.form.get("new_extract_mode", "directive").strip()
                _validate_navigation_fields(
                    new_label or new_key,
                    new_advanced,
                    new_steps,
                    new_direct_href,
                    new_sidebar,
                    new_list_link,
                    new_tab_name,
                    new_extract_mode,
                )
                settings_store.add_task(
                    {
                        "key": new_key,
                        "label": new_label,
                        "role_pattern": request.form.get("new_role_pattern", "").strip(),
                        "sidebar_item": new_sidebar,
                        "list_link": new_list_link,
                        "list_link_href": new_direct_href,
                        "tab_name": new_tab_name,
                        "navigation_steps": new_steps,
                        "use_advanced_navigation": new_advanced,
                        "document_row_selector": request.form.get("new_document_row_selector", "tr.mat-row").strip(),
                        "document_click_selector": request.form.get("new_document_click_selector", "").strip(),
                        "extract_mode": new_extract_mode,
                        "max_documents": request.form.get("new_max_documents", 20),
                        "download_subdir": request.form.get("new_download_subdir", "").strip(),
                        "sheet_name": request.form.get("new_sheet_name", "").strip(),
                        "title_text": request.form.get("new_title_text", "").strip(),
                        "enable_download_pdf": request.form.get("new_enable_download_pdf") == "on",
                        "enable_finish": request.form.get("new_enable_finish") == "on",
                    }
                )
                settings_store.reload_config(base_config)
                new_cfg = settings_store.build_effective_config(base_config)
                excel_log.ensure_all_sheets(new_cfg.EXCEL_FILE, new_cfg.TASKS)
                return redirect(url_for("settings_page", saved=1, message="Đã tạo tác vụ và sheet Excel mới."))
        except Exception as e:
            return redirect(url_for("settings_page", error=str(e)))

        doffice_url = request.form.get("doffice_url", "").strip()
        common_updates = {
            "STOP_WHEN_DUPLICATE_FOUND": request.form.get("stop_when_duplicate_found") == "on",
            "DUPLICATE_CHECK_MODE": request.form.get("duplicate_check_mode", effective_cfg.DUPLICATE_CHECK_MODE),
            "ENABLE_TELEGRAM_NOTIFY": request.form.get("enable_telegram_notify") == "on",
            "TELEGRAM_BOT_TOKEN": request.form.get("telegram_bot_token", effective_cfg.TELEGRAM_BOT_TOKEN).strip(),
            "TELEGRAM_CHAT_ID": request.form.get("telegram_chat_id", effective_cfg.TELEGRAM_CHAT_ID).strip(),
            "TELEGRAM_NOTIFY_ONLY_IF_NEW": request.form.get("telegram_notify_only_if_new") == "on",
            "DISPLAY_BASE_DIR_OVERRIDE": request.form.get(
                "display_base_dir_override", effective_cfg.DISPLAY_BASE_DIR_OVERRIDE
            ).strip(),
            "DISPLAY_BASE_URL_OVERRIDE": request.form.get(
                "display_base_url_override", effective_cfg.DISPLAY_BASE_URL_OVERRIDE
            ).strip(),
        }
        if doffice_url:
            common_updates["DOFFICE_URL"] = doffice_url
        try:
            common_updates["SLOW_MO_MS"] = int(request.form.get("slow_mo_ms", effective_cfg.SLOW_MO_MS))
        except (TypeError, ValueError):
            common_updates["SLOW_MO_MS"] = effective_cfg.SLOW_MO_MS

        error = None
        try:
            task_updates = {}
            for key, eff_task in effective_cfg.TASKS.items():
                try:
                    max_docs = int(request.form.get(f"{key}_max_documents") or eff_task.max_documents)
                except (TypeError, ValueError):
                    max_docs = eff_task.max_documents
                sidebar_item = request.form.get(f"{key}_sidebar_item", eff_task.sidebar_item).strip()
                list_link = request.form.get(f"{key}_list_link", eff_task.list_link).strip()
                direct_href = request.form.get(f"{key}_list_link_href", eff_task.list_link_href).strip()
                tab_name = request.form.get(f"{key}_tab_name", eff_task.tab_name or "").strip()
                extract_mode = request.form.get(f"{key}_extract_mode", eff_task.extract_mode).strip()
                use_advanced_navigation = request.form.get(f"{key}_use_advanced_navigation") == "on"
                navigation_steps = _parse_navigation_steps(
                    request.form.get(f"{key}_navigation_steps", json.dumps(eff_task.navigation_steps, ensure_ascii=False)),
                    eff_task.label,
                )
                _validate_navigation_fields(
                    eff_task.label,
                    use_advanced_navigation,
                    navigation_steps,
                    direct_href,
                    sidebar_item,
                    list_link,
                    tab_name,
                    extract_mode,
                )
                task_updates[key] = {
                    "label": request.form.get(f"{key}_label", eff_task.label).strip(),
                    "enabled": request.form.get(f"{key}_enabled") == "on",
                    "role_pattern": request.form.get(f"{key}_role_pattern", eff_task.role_pattern).strip(),
                    "sidebar_item": sidebar_item,
                    "list_link": list_link,
                    "list_link_href": direct_href,
                    "tab_name": tab_name or None,
                    "navigation_steps": navigation_steps,
                    "use_advanced_navigation": use_advanced_navigation,
                    "document_row_selector": request.form.get(
                        f"{key}_document_row_selector", eff_task.document_row_selector
                    ).strip(),
                    "document_click_selector": request.form.get(
                        f"{key}_document_click_selector", eff_task.document_click_selector
                    ).strip(),
                    "extract_mode": extract_mode,
                    "max_documents": max(1, max_docs),
                    "enable_finish": request.form.get(f"{key}_enable_finish") == "on",
                    "enable_download_pdf": request.form.get(f"{key}_enable_download_pdf") == "on",
                    "ask_confirm_before_finish": request.form.get(f"{key}_ask_confirm_before_finish") == "on",
                    "download_subdir": request.form.get(f"{key}_download_subdir", eff_task.download_subdir).strip(),
                    "sheet_name": request.form.get(f"{key}_sheet_name", eff_task.sheet_name).strip(),
                    "title_text": request.form.get(f"{key}_title_text", eff_task.title_text).strip(),
                }
            settings_store.update_common_fields(common_updates)
            for key, updates in task_updates.items():
                settings_store.update_task_fields(key, updates)
            settings_store.reload_config(base_config)
            refreshed_cfg = settings_store.build_effective_config(base_config)
            excel_log.ensure_all_sheets(refreshed_cfg.EXCEL_FILE, refreshed_cfg.TASKS)
        except Exception as e:
            error = str(e)

        if error:
            return redirect(url_for("settings_page", error=error))
        return redirect(url_for("settings_page", saved=1))

    tasks = _ordered_tasks(effective_cfg)
    common = {
        "DOFFICE_URL": effective_cfg.DOFFICE_URL,
        "STOP_WHEN_DUPLICATE_FOUND": effective_cfg.STOP_WHEN_DUPLICATE_FOUND,
        "DUPLICATE_CHECK_MODE": effective_cfg.DUPLICATE_CHECK_MODE,
        "SLOW_MO_MS": effective_cfg.SLOW_MO_MS,
        "ENABLE_TELEGRAM_NOTIFY": effective_cfg.ENABLE_TELEGRAM_NOTIFY,
        "TELEGRAM_BOT_TOKEN": effective_cfg.TELEGRAM_BOT_TOKEN,
        "TELEGRAM_CHAT_ID": effective_cfg.TELEGRAM_CHAT_ID,
        "TELEGRAM_NOTIFY_ONLY_IF_NEW": effective_cfg.TELEGRAM_NOTIFY_ONLY_IF_NEW,
        "DISPLAY_BASE_DIR": effective_cfg.DISPLAY_BASE_DIR,
        "DISPLAY_BASE_URL": effective_cfg.DISPLAY_BASE_URL,
        "DISPLAY_BASE_DIR_OVERRIDE": effective_cfg.DISPLAY_BASE_DIR_OVERRIDE,
        "DISPLAY_BASE_URL_OVERRIDE": effective_cfg.DISPLAY_BASE_URL_OVERRIDE,
    }
    saved = request.args.get("saved") == "1"
    message = request.args.get("message", "")
    error = request.args.get("error")
    return render_template(
        "settings.html",
        tasks=tasks,
        common=common,
        saved=saved,
        message=message,
        error=error,
        auth=_auth_state_info(effective_cfg),
        login_status=login_manager.status(),
        route_choices=DOFFICE_ROUTE_CHOICES,
        # Nut "Dang nhap lai" can mo Chromium THAT tren may dang chay server
        # nay - tren Linux/Docker (NAS/Pi) khong co man hinh (DISPLAY) se luon
        # loi, xem do_auto/login_flow.py. Bao truoc thay vi de bam xong moi biet.
        can_login_here=_can_login_interactively(),
    )


@app.route("/scheduler", methods=["GET", "POST"])
def scheduler_page():
    if request.method == "POST":
        t1 = request.form.get("time1", "").strip()
        t2 = request.form.get("time2", "").strip()
        t3 = request.form.get("time3", "").strip()
        times = [t for t in (t1, t2, t3) if t]

        if not times:
            ok, message = scheduler_mod.remove_schedule()
        else:
            ok, message = scheduler_mod.apply_schedule(PROJECT_DIR, times)

        return render_template(
            "scheduler.html", times=times, saved=True, ok=ok, message=message, is_windows=sys.platform == "win32"
        )

    current_times = scheduler_mod.get_current_times()
    return render_template(
        "scheduler.html", times=current_times, saved=False, ok=None, message="", is_windows=sys.platform == "win32"
    )


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8877, threaded=True, debug=False)
